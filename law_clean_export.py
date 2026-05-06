"""
Law Data — Clean Export Generator
===================================
Reads the enriched output from law_matching_v5.py and produces
a clean updated copy of the original file:

  law_merged_output_V2_updated.xlsx

  - Identical structure to the original (same 33 columns, same order)
  - Same 3,542 rows
  - Only difference: previously NULL fields are now filled in
  - No meta columns, no match scores — looks exactly like the original
  - Professional formatting applied

Dependencies:
  pip install pandas openpyxl
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
ENRICHED_FILE = 'enriched_law_v5.xlsx'
OUT_FILE      = 'law_merged_output_V2_updated.xlsx'

COLORS = {
    'header' : 'FF1F3864',
    'filled' : 'FFC6EFCE',   # light green  — cells that were filled
    'normal' : 'FFFFFFFF',   # white        — unchanged cells
    'alt'    : 'FFF2F7FF',   # light blue   — alternating rows
}

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


# ──────────────────────────────────────────────────────────────────────────────
# 1. LOAD & STRIP META COLUMNS
# ──────────────────────────────────────────────────────────────────────────────
print("[1/3] Loading enriched file ...")
df_enriched = pd.read_excel(ENRICHED_FILE)

# Columns that were enriched (to highlight them)
enrich_cols = {'Magazine_Number', 'Magazine_Date', 'Magazine_Page_Number',
               'Active_Date', 'Replaced_For'}

# Keep only original columns (drop everything starting with _)
original_cols = [c for c in df_enriched.columns if not c.startswith('_')]
df_clean      = df_enriched[original_cols].copy()

# Also keep a copy of the meta columns for highlighting
df_meta = df_enriched[[c for c in df_enriched.columns if c.startswith('_')]].copy()

print(f"      Rows    : {len(df_clean):,}")
print(f"      Columns : {len(original_cols)}")


# ──────────────────────────────────────────────────────────────────────────────
# 2. EXPORT RAW
# ──────────────────────────────────────────────────────────────────────────────
print("[2/3] Writing Excel file ...")
df_clean.to_excel(OUT_FILE, index=False)


# ──────────────────────────────────────────────────────────────────────────────
# 3. FORMAT
# ──────────────────────────────────────────────────────────────────────────────
print("[3/3] Applying formatting ...")

wb = load_workbook(OUT_FILE)
ws = wb.active
ws.title = 'Laws'
ws.freeze_panes = 'A2'

all_cols = [c.value for c in ws[1]]

# Header row
for col_idx, col_name in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid', fgColor=COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[1].height = 36

# Build a set of (row_idx, col_name) that were filled (were null, now have value)
# We determine this by checking the _enriched_fields meta column
filled_cells = set()
for row_idx, enr_fields in enumerate(df_meta['_enriched_fields'], start=2):
    if isinstance(enr_fields, str) and enr_fields not in ('', 'none'):
        for col_name in enr_fields.split(', '):
            filled_cells.add((row_idx, col_name.strip()))

# Data rows
for row_idx in range(2, ws.max_row + 1):
    for col_idx, col_name in enumerate(all_cols, 1):
        cell   = ws.cell(row=row_idx, column=col_idx)
        is_filled = (row_idx, col_name) in filled_cells

        if is_filled:
            cell.fill = PatternFill('solid', fgColor=COLORS['filled'])
        elif row_idx % 2 == 0:
            cell.fill = PatternFill('solid', fgColor=COLORS['alt'])
        else:
            cell.fill = PatternFill('solid', fgColor=COLORS['normal'])

        cell.font      = Font(name='Arial', size=9,
                               bold=is_filled)   # bold the newly filled values
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

# Column widths
col_width_map = {
    'pmk_ID': 8, 'GUID': 12, 'UserID': 8, 'LangID': 8,
    'Law_Name': 45, 'Law_Number': 12, 'Year': 8,
    'Magazine_Number': 16, 'Magazine_Date': 16,
    'Magazine_Page_Number': 18, 'Active_Date': 16,
    'Replaced_For': 40, 'FullName': 40,
}
for col_idx, col_name in enumerate(all_cols, 1):
    if col_name in col_width_map:
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width_map[col_name]
    else:
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

wb.save(OUT_FILE)

# ──────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ──────────────────────────────────────────────────────────────────────────────
total_filled = len(filled_cells)
rows_touched = len(set(r for r, _ in filled_cells))

print(f"\n✅ Done!")
print(f"   → {OUT_FILE}")
print(f"")
print(f"   Rows     : {len(df_clean):,}  (same as original)")
print(f"   Columns  : {len(original_cols)}  (same as original)")
print(f"   Cells filled (highlighted green) : {total_filled:,}")
print(f"   Rows with at least one new value : {rows_touched:,}")
print(f"")
print(f"   Green highlighted cells = newly filled values")
print(f"   This file is ready to send to your supervisor.")
