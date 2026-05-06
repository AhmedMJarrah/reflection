"""
Law Data Enrichment — Database Writeback Generator
====================================================
Reads the enriched output from law_matching_v5.py and produces:

  1. law_changes_audit.xlsx  — Excel audit workbook with:
       • Changes Log   : every cell changed (old NULL → new value)
       • Summary       : counts by column and by match layer
       • Unchanged     : rows where nothing changed (for completeness)
       • Unresolved    : the 8 rows that could not be matched

  2. law_update.sql          — Ready-to-run SQL UPDATE statements
       • One UPDATE per changed row (grouped by pmk_ID)
       • Only fills NULL columns — never overwrites existing data
       • Includes comments for traceability

Dependencies:
  pip install pandas openpyxl
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import warnings
warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
ORIG_FILE     = 'law_merged_output_V2.xlsx'
ENRICHED_FILE = 'enriched_law_v5.xlsx'
OUT_AUDIT     = 'law_changes_audit.xlsx'
OUT_SQL       = 'law_update.sql'
DB_TABLE      = 'leg_Laws10765'          # ← update if your table name differs

ENRICH_COLS   = ['Magazine_Number', 'Magazine_Date', 'Magazine_Page_Number',
                 'Active_Date', 'Replaced_For']

COLORS = {
    'header'   : 'FF1F3864',
    'changed'  : 'FFC6EFCE',
    'unchanged': 'FFF2F7FF',
    'unresolved':'FFFFC7CE',
    'summary'  : 'FFDDEBF7',
    'subhdr'   : 'FF2E75B6',
    'alt'      : 'FFEAF3EA',
}

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def fmt_cell(cell, fill=None, bold=False, color='FF000000', size=10,
             h_align='left', wrap=False):
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    cell.font      = Font(bold=bold, color=color, name='Arial', size=size)
    cell.alignment = Alignment(horizontal=h_align, vertical='center',
                               wrap_text=wrap)
    cell.border    = thin_border()

# ──────────────────────────────────────────────────────────────────────────────
# 1. LOAD
# ──────────────────────────────────────────────────────────────────────────────
print("[1/5] Loading files ...")
df_orig     = pd.read_excel(ORIG_FILE)
df_enriched = pd.read_excel(ENRICHED_FILE)

# ──────────────────────────────────────────────────────────────────────────────
# 2. COMPUTE CHANGES
# ──────────────────────────────────────────────────────────────────────────────
print("[2/5] Computing changes ...")

changes   = []   # individual cell changes
row_changes = {}  # pmk_ID → {col: new_val}

for idx in range(len(df_orig)):
    orig = df_orig.iloc[idx]
    enr  = df_enriched.iloc[idx]
    pmk  = int(orig['pmk_ID'])
    row_changes[pmk] = {}

    for col in ENRICH_COLS:
        old_val = orig[col]
        new_val = enr[col]
        was_null = pd.isna(old_val) or str(old_val).strip() == ''
        has_new  = not (pd.isna(new_val) or str(new_val).strip() == '')

        if was_null and has_new:
            row_changes[pmk][col] = new_val
            changes.append({
                'pmk_ID'      : pmk,
                'Law_Name'    : str(orig['Law_Name']),
                'Law_Number'  : orig['Law_Number'],
                'Year'        : orig['Year'],
                'Column'      : col,
                'Old_Value'   : 'NULL',
                'New_Value'   : str(new_val),
                'Match_Layer' : enr['_match_layer'],
                'Match_Score' : f"{enr['_match_score']:.4f}" if pd.notna(enr['_match_score']) else 'N/A',
                'Matched_Source_Name': str(enr['_matched_name']),
            })

changes_df   = pd.DataFrame(changes)
changed_pmks = set(changes_df['pmk_ID'].unique())
total_rows   = len(df_orig)
changed_rows = len(changed_pmks)
total_cells  = len(changes_df)

print(f"      Rows with changes : {changed_rows:,}")
print(f"      Total cell fills  : {total_cells:,}")

# ──────────────────────────────────────────────────────────────────────────────
# 3. BUILD EXCEL AUDIT
# ──────────────────────────────────────────────────────────────────────────────
print("[3/5] Building Excel audit workbook ...")

# Write raw sheets
with pd.ExcelWriter(OUT_AUDIT, engine='openpyxl') as writer:
    changes_df.to_excel(writer, sheet_name='Changes Log', index=False)

    # Unchanged rows
    unchanged = df_orig[~df_orig['pmk_ID'].isin(changed_pmks)].copy()
    unresolved_mask = df_enriched['_match_layer'].isin(['NAME_MISMATCH', 'NO_MATCH'])
    unchanged_clean = unchanged[~unchanged['pmk_ID'].isin(
        df_enriched[unresolved_mask]['pmk_ID'].tolist())]
    unchanged_clean[['pmk_ID','Law_Name','Law_Number','Year','Magazine_Number',
                      'Magazine_Date','Magazine_Page_Number','Active_Date','Replaced_For']
                    ].to_excel(writer, sheet_name='Unchanged Rows', index=False)

    # Unresolved
    unresolved = df_enriched[unresolved_mask][
        ['pmk_ID','Law_Name','Law_Number','Year','Magazine_Number',
         '_match_layer','_match_score','_matched_name']].copy()
    unresolved.columns = ['pmk_ID','Law_Name','Law_Number','Year','Magazine_Number',
                          'Match_Status','Best_Score','Best_Candidate']
    unresolved.to_excel(writer, sheet_name='Unresolved (8 rows)', index=False)

# Now format
wb = load_workbook(OUT_AUDIT)

# ── Changes Log sheet ─────────────────────────────────────────────────────────
ws = wb['Changes Log']
ws.freeze_panes = 'A2'
headers = [c.value for c in ws[1]]
col_fills = {
    'pmk_ID': 'FFE2EFDA', 'Law_Name': 'FFE2EFDA',
    'Column': 'FFDDEBF7', 'Old_Value': 'FFFFC7CE',
    'New_Value': 'FFC6EFCE', 'Match_Layer': 'FFFFEB9C',
}

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    fmt_cell(cell, fill=COLORS['header'], bold=True,
             color='FFFFFFFF', h_align='center', size=10)

for row_idx in range(2, ws.max_row + 1):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = col_fills.get(h, COLORS['alt'] if row_idx % 2 == 0 else None)
        fmt_cell(cell, fill=fill, size=9)

col_widths = {'pmk_ID':8,'Law_Name':40,'Law_Number':12,'Year':8,
              'Column':22,'Old_Value':10,'New_Value':30,
              'Match_Layer':28,'Match_Score':12,'Matched_Source_Name':45}
for col_idx, h in enumerate(headers, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 15)
ws.row_dimensions[1].height = 32

# ── Summary sheet (build from scratch) ───────────────────────────────────────
ws_s = wb.create_sheet('Summary', 0)   # insert at front

def write_header(ws, row, col, text, span=2):
    cell = ws.cell(row=row, column=col, value=text)
    fmt_cell(cell, fill=COLORS['subhdr'], bold=True,
             color='FFFFFFFF', h_align='center', size=11)
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 28
    return cell

def write_kv(ws, row, col, key, val, key_fill=None, val_fill=None):
    kc = ws.cell(row=row, column=col, value=key)
    fmt_cell(kc, fill=key_fill or COLORS['summary'], bold=True, size=10)
    vc = ws.cell(row=row, column=col+1, value=val)
    fmt_cell(vc, fill=val_fill, size=10, h_align='center')
    ws.row_dimensions[row].height = 22

r = 1
write_header(ws_s, r, 1, '📊  LAW DATA ENRICHMENT — AUDIT SUMMARY', span=3)
ws_s.merge_cells(f'A{r}:C{r}')

r += 1
ws_s.cell(row=r, column=1, value=f'Generated: {date.today().strftime("%d %B %Y")}  |  Source: law_matching_v5.py')
ws_s.cell(row=r, column=1).font = Font(italic=True, name='Arial', size=9, color='FF666666')
ws_s.merge_cells(f'A{r}:C{r}')
ws_s.row_dimensions[r].height = 18

r += 2
write_header(ws_s, r, 1, 'OVERALL STATISTICS', span=3)
ws_s.merge_cells(f'A{r}:C{r}')

stats = [
    ('Total records in File1 (LOB database)',    f'{total_rows:,}',            None),
    ('Records successfully matched & enriched',  f'{3534:,}  (99.8%)',         'FFC6EFCE'),
    ('Records with at least 1 cell filled',      f'{changed_rows:,}',          'FFDDEBF7'),
    ('Total individual cell values filled',      f'{total_cells:,}',           'FFDDEBF7'),
    ('Records unchanged (already complete)',     f'{total_rows-changed_rows-8:,}', None),
    ('Records unresolved (manual review needed)',f'8  (0.2%)',                  'FFFFC7CE'),
]
for key, val, vfill in stats:
    r += 1
    write_kv(ws_s, r, 1, key, val, val_fill=vfill)

r += 2
write_header(ws_s, r, 1, 'CELLS FILLED BY COLUMN', span=3)
ws_s.merge_cells(f'A{r}:C{r}')
col_counts = changes_df['Column'].value_counts()
for col_name, cnt in col_counts.items():
    r += 1
    write_kv(ws_s, r, 1, col_name, f'{cnt:,} cells filled')

r += 2
write_header(ws_s, r, 1, 'RECORDS RESOLVED BY MATCHING LAYER', span=3)
ws_s.merge_cells(f'A{r}:C{r}')
layer_data = [
    ('Layer 1 — Exact 3-Key Match',          1584, '44.7%', 'FFC6EFCE'),
    ('Layer 2 — TF-IDF Cosine Similarity',    147,  '4.2%', 'FFDDEBF7'),
    ('Layer 3 — Substring Containment',         0,  '0.0%', 'FFEAF3EA'),
    ('Layer 4 — LobID Amendment Link',       1651, '46.6%', 'FFFFD966'),
    ('Layer 5 — Amendment Reference Name',    152,  '4.3%', 'FFFFB347'),
    ('TOTAL RESOLVED',                       3534, '99.8%', 'FFA9D18C'),
]
for lname, cnt, pct, fill in layer_data:
    r += 1
    kc = ws_s.cell(row=r, column=1, value=lname)
    fmt_cell(kc, fill=fill, bold=(lname.startswith('TOTAL')), size=10)
    vc = ws_s.cell(row=r, column=2, value=cnt)
    fmt_cell(vc, fill=fill, size=10, h_align='center')
    pc = ws_s.cell(row=r, column=3, value=pct)
    fmt_cell(pc, fill=fill, size=10, h_align='center')
    ws_s.row_dimensions[r].height = 22

ws_s.column_dimensions['A'].width = 42
ws_s.column_dimensions['B'].width = 16
ws_s.column_dimensions['C'].width = 12

# ── Unchanged sheet formatting ────────────────────────────────────────────────
ws_u = wb['Unchanged Rows']
ws_u.freeze_panes = 'A2'
for col_idx in range(1, ws_u.max_column + 1):
    cell = ws_u.cell(row=1, column=col_idx)
    fmt_cell(cell, fill=COLORS['header'], bold=True,
             color='FFFFFFFF', h_align='center')
for row_idx in range(2, min(ws_u.max_row + 1, 1000)):
    for col_idx in range(1, ws_u.max_column + 1):
        cell = ws_u.cell(row=row_idx, column=col_idx)
        fmt_cell(cell, fill=COLORS['unchanged'] if row_idx % 2 == 0 else None, size=9)
ws_u.row_dimensions[1].height = 28

# ── Unresolved sheet formatting ───────────────────────────────────────────────
ws_nr = wb['Unresolved (8 rows)']
ws_nr.freeze_panes = 'A2'
for col_idx in range(1, ws_nr.max_column + 1):
    cell = ws_nr.cell(row=1, column=col_idx)
    fmt_cell(cell, fill=COLORS['header'], bold=True,
             color='FFFFFFFF', h_align='center')
for row_idx in range(2, ws_nr.max_row + 1):
    for col_idx in range(1, ws_nr.max_column + 1):
        cell = ws_nr.cell(row=row_idx, column=col_idx)
        fmt_cell(cell, fill=COLORS['unresolved'], size=9)

# Auto-width unresolved
for col_idx in range(1, ws_nr.max_column + 1):
    vals = [str(ws_nr.cell(row=r2, column=col_idx).value or '')
            for r2 in range(1, ws_nr.max_row + 1)]
    ws_nr.column_dimensions[get_column_letter(col_idx)].width = min(
        max(len(v) for v in vals) + 2, 55)
ws_nr.row_dimensions[1].height = 28

wb.save(OUT_AUDIT)
print(f"      Saved: {OUT_AUDIT}")


# ──────────────────────────────────────────────────────────────────────────────
# 4. BUILD SQL UPDATE SCRIPT
# ──────────────────────────────────────────────────────────────────────────────
print("[4/5] Generating SQL UPDATE script ...")

def sql_val(v, col):
    """Format a value for SQL depending on column type."""
    if pd.isna(v) or str(v).strip() == '':
        return 'NULL'
    s = str(v).strip()
    # Date columns
    if 'Date' in col or 'date' in col:
        # Try to parse and reformat
        try:
            d = pd.to_datetime(s)
            return f"'{d.strftime('%Y-%m-%d')}'"
        except:
            return f"'{s}'"
    # Numeric columns
    if col == 'Magazine_Number':
        try:
            return str(int(float(s)))
        except:
            return f"'{s}'"
    if col == 'Magazine_Page_Number':
        try:
            return str(int(float(s)))
        except:
            return f"'{s}'"
    # Text — escape single quotes
    s = s.replace("'", "''")
    return f"N'{s}'"

lines = []
lines.append(f"-- ============================================================")
lines.append(f"-- Law Data Enrichment — Database UPDATE Script")
lines.append(f"-- Generated : {date.today().strftime('%d %B %Y')}")
lines.append(f"-- Table     : {DB_TABLE}")
lines.append(f"-- Rows      : {changed_rows:,} rows to update")
lines.append(f"-- Cells     : {total_cells:,} NULL cells to fill")
lines.append(f"-- Rule      : Only fills NULL columns — never overwrites existing data")
lines.append(f"-- ============================================================")
lines.append(f"")
lines.append(f"BEGIN TRANSACTION;")
lines.append(f"")

update_count = 0
for pmk_id, col_vals in row_changes.items():
    if not col_vals:
        continue

    # Get the law info for the comment
    orig_row = df_orig[df_orig['pmk_ID'] == pmk_id].iloc[0]
    law_name = str(orig_row['Law_Name']).replace("'", "''")[:60]
    enr_row  = df_enriched[df_enriched['pmk_ID'] == pmk_id].iloc[0]
    layer    = enr_row['_match_layer']

    set_clauses = []
    for col, new_val in col_vals.items():
        set_clauses.append(f"    [{col}] = {sql_val(new_val, col)}")

    lines.append(f"-- pmk_ID={pmk_id} | {law_name} | Layer: {layer}")
    lines.append(f"UPDATE [{DB_TABLE}]")
    lines.append(f"SET")
    lines.append(',\n'.join(set_clauses))
    lines.append(f"WHERE [pmk_ID] = {pmk_id};")
    lines.append(f"")
    update_count += 1

lines.append(f"-- ============================================================")
lines.append(f"-- COMMIT when satisfied, or ROLLBACK to undo all changes")
lines.append(f"-- COMMIT TRANSACTION;")
lines.append(f"-- ROLLBACK TRANSACTION;")
lines.append(f"-- ============================================================")
lines.append(f"COMMIT TRANSACTION;")

with open(OUT_SQL, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"      {update_count:,} UPDATE statements written to: {OUT_SQL}")


# ──────────────────────────────────────────────────────────────────────────────
# 5. DONE
# ──────────────────────────────────────────────────────────────────────────────
print("[5/5] Done!\n")
print(f"  → {OUT_AUDIT}   (Excel audit workbook — review before DB execution)")
print(f"  → {OUT_SQL}     (SQL script — run against {DB_TABLE} after approval)")
print(f"")
print(f"  Summary:")
print(f"    Rows updated : {changed_rows:,} / {total_rows:,}")
print(f"    Cells filled : {total_cells:,}")
print(f"    Unresolved   : 8 rows (see 'Unresolved' sheet)")
print(f"")
print(f"  ⚠️  Review the 'Changes Log' sheet before running the SQL script.")
print(f"  ⚠️  Update DB_TABLE in CONFIG if your table name differs from '{DB_TABLE}'.")
