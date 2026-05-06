"""
Final Clean Export — law_merged_output_V2_final.xlsx
=====================================================
Reads phase2_expanded.xlsx and produces a clean,
professionally formatted file ready to send to your supervisor.

Visual guide:
  🟩 Bright green cell  = field enriched in Phase 1 (was NULL, now filled)
  🟦 Light blue row     = new law added in Phase 2 (from File2)
  ⬜ White / gray row   = original File1 record (unchanged)

New columns visible:
  end_date  → when the law ceased to be in effect
  Source    → "Original" or "Added from File2"

Run: python law_final_export.py
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
ORIG_FILE     = 'law_merged_output_V2.xlsx'       # original for comparison
EXPANDED_FILE = 'phase2_expanded.xlsx'             # Phase 2 output
OUT_FILE      = 'law_merged_output_V2_final_2.xlsx'  # what you send supervisor

ENRICH_COLS   = ['Magazine_Number','Magazine_Date','Magazine_Page_Number',
                 'Active_Date','Replaced_For']     # Phase 1 enriched fields

COLORS = {
    'header'  : 'FF1F3864',   # dark navy    — header row
    'new_col' : 'FF2E75B6',   # blue         — new column headers
    'enriched': 'FF92D050',   # bright green — Phase 1 filled cell
    'new_row' : 'FFBDD7EE',   # light blue   — Phase 2 new law row
    'alt'     : 'FFF2F7FF',   # very light   — alternating original rows
    'white'   : 'FFFFFFFF',
}

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

# Arabic text columns — set RTL reading order so Excel Find & Replace works
ARABIC_COLS = {'Law_Name','FullName','Replaced_For','Magazine','Status'}

def fmt(cell, fill=None, bold=False, color='FF000000', size=10,
        align='left', col_name=None):
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    cell.font   = Font(bold=bold, color=color, name='Arial', size=size)
    # RTL reading order for Arabic columns so Excel Find works correctly
    reading_order = 2.0 if col_name in ARABIC_COLS else 0.0
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               readingOrder=reading_order)
    cell.border = thin_border()

# ─────────────────────────────────────────────────────────────────────────────
# 1. LOAD
# ─────────────────────────────────────────────────────────────────────────────
print("[1/4] Loading files ...")
df_orig = pd.read_excel(ORIG_FILE)
df_exp  = pd.read_excel(EXPANDED_FILE)

orig_count = len(df_orig)   # first N rows = original File1 records
total_rows = len(df_exp)
new_count  = total_rows - orig_count

print(f"      Original rows : {orig_count:,}")
print(f"      New rows      : {new_count:,}")
print(f"      Total         : {total_rows:,}")

# ─────────────────────────────────────────────────────────────────────────────
# 2. BUILD CHANGED CELLS SET (Phase 1 enriched cells)
# ─────────────────────────────────────────────────────────────────────────────
print("[2/4] Computing enriched cells ...")
enriched_cells = set()   # (excel_row_idx, col_name)

for idx in range(orig_count):
    orig_row = df_orig.iloc[idx]
    exp_row  = df_exp.iloc[idx]
    for col in ENRICH_COLS:
        old = orig_row[col] if col in df_orig.columns else None
        new = exp_row[col]  if col in df_exp.columns  else None
        was_null = pd.isna(old) or str(old).strip() == ''
        has_new  = not (pd.isna(new) or str(new).strip() == '')
        if was_null and has_new:
            enriched_cells.add((idx + 2, col))   # +2 for header + 0-index

# Also mark end_date cells filled for original rows
if 'end_date' in df_exp.columns:
    for idx in range(orig_count):
        val = df_exp.iloc[idx]['end_date']
        if pd.notna(val) and str(val).strip() not in ('','nan','0'):
            enriched_cells.add((idx + 2, 'end_date'))

total_enriched = len(enriched_cells)
print(f"      Enriched cells : {total_enriched:,}")

# ─────────────────────────────────────────────────────────────────────────────
# 3. EXPORT RAW
# ─────────────────────────────────────────────────────────────────────────────
print("[3/4] Writing and formatting ...")
df_exp.to_excel(OUT_FILE, index=False)

# ─────────────────────────────────────────────────────────────────────────────
# 4. FORMAT
# ─────────────────────────────────────────────────────────────────────────────
NEW_COLS = {'end_date', 'Source'}
col_widths = {
    'Law_Name':45, 'Law_Number':12, 'Year':8,
    'Magazine_Number':16, 'Magazine_Date':16,
    'Magazine_Page_Number':18, 'Active_Date':16,
    'end_date':16, 'Replaced_For':36,
    'Status':14, 'Source':20, 'FullName':38,
    'pmk_ID':8, 'LobID':10, 'GUID':12,
}

wb = load_workbook(OUT_FILE)
ws = wb.active
ws.title  = 'Laws — Final'
ws.freeze_panes = 'A2'

cols_ws = [c.value for c in ws[1]]

# ── Header row ────────────────────────────────────────────────────────────────
for ci, cn in enumerate(cols_ws, 1):
    cell  = ws.cell(row=1, column=ci)
    hfill = COLORS['new_col'] if cn in NEW_COLS else COLORS['header']
    fmt(cell, fill=hfill, bold=True, color='FFFFFFFF', align='center', size=10)
ws.row_dimensions[1].height = 38

# ── Data rows ─────────────────────────────────────────────────────────────────
for ri in range(2, ws.max_row + 1):
    is_new_law = ri > orig_count + 1   # beyond original rows

    for ci, cn in enumerate(cols_ws, 1):
        cell = ws.cell(row=ri, column=ci)

        if is_new_law:
            fill  = COLORS['new_row']
            bold  = cn in ('Law_Name', 'Source')
            color = 'FF1F3864' if bold else 'FF000000'
        elif (ri, cn) in enriched_cells:
            fill  = COLORS['enriched']
            bold  = True
            color = 'FF375623'
        elif ri % 2 == 0:
            fill  = COLORS['alt']
            bold  = False
            color = 'FF000000'
        else:
            fill  = COLORS['white']
            bold  = False
            color = 'FF000000'

        cell.fill  = PatternFill('solid', fgColor=fill)
        cell.font  = Font(name='Arial', size=9, bold=bold, color=color)
        ro = 2.0 if cn in ARABIC_COLS else 0.0
        cell.alignment = Alignment(vertical='center', readingOrder=ro)
        cell.border = thin_border()
    ws.row_dimensions[ri].height = 20

# ── Column widths ─────────────────────────────────────────────────────────────
for ci, cn in enumerate(cols_ws, 1):
    ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 13)

# ── Legend sheet ──────────────────────────────────────────────────────────────
ws_leg = wb.create_sheet('Legend')
legend_rows = [
    ('Color / Style',              'Meaning'),
    ('🟩 Bright green cell',        'Field enriched in Phase 1 — was empty, now filled from matched source'),
    ('🟦 Light blue row',           'New law added in Phase 2 — did not exist in original File1'),
    ('⬜ White / gray row',          'Original File1 record — unchanged'),
    ('"Source" column = Original',  'Law existed in the original law_merged_output_V2 database'),
    ('"Source" column = Added from File2', 'Law was newly appended from cleaned_v03_merged_updated'),
    ('"end_date" column',           'Date when the law ceased to be in effect (new column added in Phase 2)'),
]
legend_fills = [
    COLORS['header'], COLORS['enriched'], COLORS['new_row'],
    COLORS['white'], COLORS['alt'], COLORS['alt'], COLORS['new_col'],
]
legend_colors = ['FFFFFFFF','FF375623','FF1F3864',
                 'FF000000','FF000000','FF000000','FFFFFFFF']

for r, (label, desc) in enumerate(legend_rows, 1):
    for c, val in enumerate([label, desc], 1):
        cell = ws_leg.cell(row=r, column=c, value=val)
        cell.fill      = PatternFill('solid', fgColor=legend_fills[r-1])
        cell.font      = Font(bold=(r==1), color=legend_colors[r-1],
                               name='Arial', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border    = thin_border()
    ws_leg.row_dimensions[r].height = 26

ws_leg.column_dimensions['A'].width = 36
ws_leg.column_dimensions['B'].width = 72

wb.save(OUT_FILE)

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print()
print("  ╔══════════════════════════════════════════════════════╗")
print("  ║         FINAL EXPORT — READY TO SEND                ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  Original File1 rows   : {orig_count:>6,}                  ║")
print(f"  ║  New laws added        : {new_count:>6,}                  ║")
print(f"  ║  Total rows            : {total_rows:>6,}                  ║")
print(f"  ║  Enriched cells (green): {total_enriched:>6,}                  ║")
print("  ╠══════════════════════════════════════════════════════╣")
print("  ║  New columns: end_date, Source                       ║")
print("  ║  Sheets: 'Laws — Final' + 'Legend'                   ║")
print("  ╚══════════════════════════════════════════════════════╝")
print()
print(f"  ✅  Ready → {OUT_FILE}")
print()
print("  Color guide:")
print("    🟩 Bright green cell  = Phase 1 enriched field")
print("    🟦 Light blue row     = Phase 2 new law")
print("    ⬜ White / gray       = original unchanged record")