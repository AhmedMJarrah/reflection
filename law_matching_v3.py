"""
Arabic Law Matching — Multi-Strategy Pipeline v3
=================================================
Matching layers (applied in order, first match wins):

  Layer 1 — EXACT KEY         : Law_Number + Year + Magazine_Number  (strictest)
  Layer 2 — TFIDF COSINE      : Law_Number + Year, cosine ≥ COSINE_THRESHOLD
  Layer 3 — SUBSTRING         : Law_Number + Year, normalized File1 name
                                 is contained inside File2 candidate name
  Layer 4 — UNMATCHED         : flagged for manual review

Each matched row records WHICH layer resolved it, giving you full transparency.

Outputs:
  enriched_law_v3.xlsx      — File1 enriched, color-coded by match layer
  mismatch_report_v3.xlsx   — Only the truly unresolvable rows

Dependencies:
  pip install pandas openpyxl scikit-learn
"""

import re
import unicodedata
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# ──────────────────────────────────────────────────────────────────────────────
# CONFIG  —  tune these without touching anything else
# ──────────────────────────────────────────────────────────────────────────────
FILE1_PATH       = 'law_merged_output_V2.xlsx'
FILE2_PATH       = 'cleaned_v03_merged_updated.xlsx'
OUT_ENRICHED     = 'enriched_law_v3.xlsx'
OUT_MISMATCH     = 'mismatch_report_v3.xlsx'

COSINE_THRESHOLD       = 0.35   # Layer 2 cut-off
SUBSTRING_MIN_LENGTH   = 6      # ignore very short names in substring check (chars)
NGRAM_RANGE            = (2, 4)

# File1 → File2 enrichment column mapping
ENRICH_MAP = {
    'Magazine_Number'     : 'magazine_number',
    'Magazine_Date'       : 'magazine_date',
    'Magazine_Page_Number': 'magazine_page',
    'Active_Date'         : 'start_date',
    'Replaced_For'        : 'replaced_for',
}

COLORS = {
    'header'     : 'FF1F3864',
    'layer1'     : 'FFC6EFCE',   # green   — exact 3-key
    'layer2'     : 'FFDDEBF7',   # blue    — TF-IDF cosine
    'layer3'     : 'FFEAF3EA',   # mint    — substring
    'mismatch'   : 'FFFFEB9C',   # amber   — name conflict
    'no_match'   : 'FFFFC7CE',   # red     — key not found
    'meta_hdr'   : 'FF2E75B6',
    'alt'        : 'FFF2F7FF',
}

LAYER_LABELS = {
    'EXACT_3KEY'  : '✅ Exact (Num+Year+Mag)',
    'TFIDF'       : '✅ TF-IDF Cosine',
    'SUBSTRING'   : '✅ Substring Containment',
    'NAME_MISMATCH': '⚠️  Name Mismatch',
    'NO_MATCH'    : '❌ No Match',
}


# ──────────────────────────────────────────────────────────────────────────────
# 1. ARABIC NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
def normalize_arabic(text: str) -> str:
    """
    Strips tashkeel, unifies Alef/Hamza/Teh-marbuta/Yeh variants,
    removes tatweel, collapses whitespace.
    """
    if not isinstance(text, str) or not text.strip():
        return ''
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[أإآٱ]', 'ا', text)
    text = re.sub(r'[ئؤ]', 'ء', text)
    text = text.replace('ة', 'ه').replace('ى', 'ي').replace('ـ', '')
    # strip parentheses and punctuation that wrap formal titles
    text = re.sub(r'[().,،؛؟!]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def to_int_str(val) -> str:
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return ''


def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


# ──────────────────────────────────────────────────────────────────────────────
# 2. LOAD & PREPARE
# ──────────────────────────────────────────────────────────────────────────────
print("[1/7] Loading files ...")
df1 = pd.read_excel(FILE1_PATH)
df2 = pd.read_excel(FILE2_PATH)
print(f"      File1: {len(df1):,} rows  |  File2: {len(df2):,} rows")

for df, num_col, year_col, mag_col in [
    (df1, 'Law_Number', 'Year',  'Magazine_Number'),
    (df2, 'leg_number', 'year',  'magazine_number'),
]:
    df['_num']  = df[num_col].apply(to_int_str)
    df['_year'] = df[year_col].apply(to_int_str)
    df['_mag']  = df[mag_col].apply(to_int_str)
    df['_key2'] = df['_num'] + '|' + df['_year']                    # 2-field key
    df['_key3'] = df['_num'] + '|' + df['_year'] + '|' + df['_mag'] # 3-field key

name_col_f1 = 'Law_Name'
name_col_f2 = 'leg_name'
df1['_norm'] = df1[name_col_f1].apply(normalize_arabic)
df2['_norm'] = df2[name_col_f2].apply(normalize_arabic)


# ──────────────────────────────────────────────────────────────────────────────
# 3. TF-IDF INDEX  (built once on all File2 names)
# ──────────────────────────────────────────────────────────────────────────────
print("[2/7] Building TF-IDF index ...")
vectorizer = TfidfVectorizer(
    analyzer    = 'char_wb',
    ngram_range = NGRAM_RANGE,
    sublinear_tf= True,
    min_df      = 1,
)
tfidf_f2 = vectorizer.fit_transform(df2['_norm'])


# ──────────────────────────────────────────────────────────────────────────────
# 4. LOOKUP TABLES
# ──────────────────────────────────────────────────────────────────────────────
df2_by_key2 = df2.groupby('_key2')   # candidates by Num+Year
df2_by_key3 = df2.set_index('_key3') # for exact 3-key lookup


# ──────────────────────────────────────────────────────────────────────────────
# 5. MULTI-LAYER MATCHING
# ──────────────────────────────────────────────────────────────────────────────
print("[3/7] Matching (3 layers) ...")

def enrich_record(record: dict, best_row: pd.Series, layer: str,
                  score: float) -> dict:
    """Fill enrichment columns from the matched File2 row."""
    record['_match_layer'] = layer
    record['_match_score'] = round(float(score), 4)
    record['_matched_leg_name'] = best_row[name_col_f2]
    enriched = []
    for f1_col, f2_col in ENRICH_MAP.items():
        val = best_row[f2_col]
        if pd.notna(val) and val != '':
            record[f1_col] = val
            enriched.append(f1_col)
    record['_enriched_fields'] = ', '.join(enriched) if enriched else 'none'
    return record


results = []

for _, row1 in df1.iterrows():
    record = row1.to_dict()
    record['_match_layer']    = 'NO_MATCH'
    record['_match_score']    = np.nan
    record['_matched_leg_name'] = ''
    record['_enriched_fields']  = ''

    key2 = row1['_key2']
    key3 = row1['_key3']
    norm1 = row1['_norm']

    # ── Layer 1: exact 3-key (Num + Year + Magazine_Number) ──────────────────
    if key3 and key3 in df2_by_key3.index:
        rows3 = df2_by_key3.loc[[key3]]
        # among these, pick best name score for transparency
        best_idx, best_score = 0, 0.0
        for i, (_, r) in enumerate(rows3.iterrows()):
            s = cosine_similarity(
                vectorizer.transform([norm1]),
                tfidf_f2[[df2.index.get_loc(r.name) if r.name in df2.index else 0]]
            )[0][0]
            if s > best_score:
                best_score, best_idx = s, i
        best_row = rows3.iloc[best_idx]
        record = enrich_record(record, best_row, 'EXACT_3KEY', best_score)
        results.append(record)
        continue

    # ── Layers 2 & 3 need candidate pool by Num+Year ─────────────────────────
    if key2 not in df2_by_key2.groups:
        results.append(record)   # NO_MATCH — key not in File2 at all
        continue

    cand_idx  = list(df2_by_key2.groups[key2])
    candidates = df2.loc[cand_idx]
    cand_vecs  = tfidf_f2[cand_idx]

    # ── Layer 2: TF-IDF cosine ───────────────────────────────────────────────
    if norm1:
        query_vec  = vectorizer.transform([norm1])
        cos_scores = cosine_similarity(query_vec, cand_vecs).flatten()
        best_pos   = cos_scores.argmax()
        best_score = cos_scores[best_pos]
        best_row   = candidates.iloc[best_pos]

        if best_score >= COSINE_THRESHOLD:
            record = enrich_record(record, best_row, 'TFIDF', best_score)
            results.append(record)
            continue

    # ── Layer 3: substring containment ───────────────────────────────────────
    # Check if File1 normalized name is wholly contained in any candidate name.
    # Also check the reverse (candidate short name inside File1 long name).
    if norm1 and len(norm1) >= SUBSTRING_MIN_LENGTH:
        sub_match = None
        sub_score = 0.0
        for i, (_, cand_row) in enumerate(candidates.iterrows()):
            norm2 = cand_row['_norm']
            if not norm2:
                continue
            # forward:  File1 inside File2
            forward  = norm1 in norm2
            # reverse:  File2 inside File1  (e.g. File2 is the short label)
            reverse  = norm2 in norm1 and len(norm2) >= SUBSTRING_MIN_LENGTH
            if forward or reverse:
                # score = character overlap ratio for ranking multiple hits
                overlap = len(norm1) / len(norm2) if forward else len(norm2) / len(norm1)
                if overlap > sub_score:
                    sub_score = overlap
                    sub_match = cand_row

        if sub_match is not None:
            record = enrich_record(record, sub_match, 'SUBSTRING', sub_score)
            results.append(record)
            continue

    # ── No layer matched ─────────────────────────────────────────────────────
    # Record the best cosine candidate for the mismatch report
    if norm1 and key2 in df2_by_key2.groups:
        query_vec  = vectorizer.transform([norm1])
        cos_scores = cosine_similarity(query_vec, cand_vecs).flatten()
        best_pos   = cos_scores.argmax()
        record['_match_score']      = round(float(cos_scores[best_pos]), 4)
        record['_matched_leg_name'] = candidates.iloc[best_pos][name_col_f2]
        record['_match_layer']      = 'NAME_MISMATCH'

    results.append(record)


result_df = pd.DataFrame(results)


# ──────────────────────────────────────────────────────────────────────────────
# 6. SUMMARY
# ──────────────────────────────────────────────────────────────────────────────
total = len(result_df)
counts = result_df['_match_layer'].value_counts().to_dict()

layer1   = counts.get('EXACT_3KEY',   0)
layer2   = counts.get('TFIDF',        0)
layer3   = counts.get('SUBSTRING',    0)
mismatch = counts.get('NAME_MISMATCH',0)
no_match = counts.get('NO_MATCH',     0)
resolved = layer1 + layer2 + layer3

print(f"\n{'='*58}")
print(f"  MATCHING SUMMARY  (threshold = {COSINE_THRESHOLD})")
print(f"{'='*58}")
print(f"  Total File1 records        :  {total:>6,}")
print(f"  ─────────────────────────────────────────")
print(f"  ✅ Layer 1 — Exact 3-key   :  {layer1:>6,}  ({100*layer1/total:.1f}%)")
print(f"  ✅ Layer 2 — TF-IDF cosine :  {layer2:>6,}  ({100*layer2/total:.1f}%)")
print(f"  ✅ Layer 3 — Substring     :  {layer3:>6,}  ({100*layer3/total:.1f}%)")
print(f"  ─────────────────────────────────────────")
print(f"  ✅ TOTAL RESOLVED          :  {resolved:>6,}  ({100*resolved/total:.1f}%)")
print(f"  ⚠️  Name mismatch           :  {mismatch:>6,}  ({100*mismatch/total:.1f}%)")
print(f"  ❌ No match (key absent)   :  {no_match:>6,}  ({100*no_match/total:.1f}%)")
print(f"{'='*58}\n")


# ──────────────────────────────────────────────────────────────────────────────
# 7. EXPORT RAW FILES
# ──────────────────────────────────────────────────────────────────────────────
print("[4/7] Exporting raw Excel files ...")

output_cols = [c for c in df1.columns if not c.startswith('_')]
meta_cols   = ['_match_layer', '_match_score', '_matched_leg_name', '_enriched_fields']

result_df[output_cols + meta_cols].to_excel(OUT_ENRICHED, index=False)

mismatch_df = result_df[
    result_df['_match_layer'].isin(['NAME_MISMATCH', 'NO_MATCH'])
][['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
   '_match_layer', '_match_score', '_matched_leg_name']].copy()
mismatch_df.columns = ['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
                        'Match_Status', 'Best_Score', 'Best_Candidate_in_File2']
mismatch_df.to_excel(OUT_MISMATCH, index=False)


# ──────────────────────────────────────────────────────────────────────────────
# 8. FORMATTING — ENRICHED FILE
# ──────────────────────────────────────────────────────────────────────────────
print("[5/7] Formatting enriched file ...")

LAYER_COLOR = {
    'EXACT_3KEY'   : COLORS['layer1'],
    'TFIDF'        : COLORS['layer2'],
    'SUBSTRING'    : COLORS['layer3'],
    'NAME_MISMATCH': COLORS['mismatch'],
    'NO_MATCH'     : COLORS['no_match'],
}

wb = load_workbook(OUT_ENRICHED)
ws = wb.active
ws.title = 'Enriched Data'

all_cols       = [c.value for c in ws[1]]
meta_col_set   = set(meta_cols)
status_col_idx = all_cols.index('_match_layer') + 1

for col_idx, col_name in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid',
                                  fgColor=COLORS['meta_hdr'] if col_name in meta_col_set
                                  else COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[1].height = 36

for row_idx in range(2, ws.max_row + 1):
    layer      = ws.cell(row=row_idx, column=status_col_idx).value
    fill_color = LAYER_COLOR.get(layer, COLORS['alt'] if row_idx % 2 == 0 else None)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

for col_idx, col_name in enumerate(all_cols, 1):
    vals  = [str(ws.cell(row=r, column=col_idx).value or '')
             for r in range(1, min(ws.max_row + 1, 60))]
    width = min(max(max(len(v) for v in vals) + 2, 12), 50)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'

# ── Summary sheet ────────────────────────────────────────────────────────────
ws_s = wb.create_sheet('Summary')

summary_rows = [
    ('Layer',                           'Count',   'Share',                     COLORS['header']),
    ('Total File1 Records',              total,    '100%',                       COLORS['header']),
    ('Layer 1 - Exact 3-Key',            layer1,  f'{100*layer1/total:.1f}%',   COLORS['layer1']),
    ('Layer 2 - TF-IDF Cosine',          layer2,  f'{100*layer2/total:.1f}%',   COLORS['layer2']),
    ('Layer 3 - Substring',              layer3,  f'{100*layer3/total:.1f}%',   COLORS['layer3']),
    ('TOTAL RESOLVED',                  resolved, f'{100*resolved/total:.1f}%', 'FFA9D18C'),
    ('Name Mismatch',                   mismatch, f'{100*mismatch/total:.1f}%', COLORS['mismatch']),
    ('No Match (key absent)',            no_match, f'{100*no_match/total:.1f}%', COLORS['no_match']),
]

for r, (label, count, share, color) in enumerate(summary_rows, 1):
    for c, val in enumerate([label, count, share], 1):
        cell = ws_s.cell(row=r, column=c, value=val)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.font      = Font(bold=(r in (1, 6)),
                               color='FFFFFFFF' if r <= 2 else 'FF000000',
                               name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border()
    ws_s.row_dimensions[r].height = 26

ws_s.column_dimensions['A'].width = 34
ws_s.column_dimensions['B'].width = 12
ws_s.column_dimensions['C'].width = 12

note = (f'Config: COSINE_THRESHOLD={COSINE_THRESHOLD}  |  '
        f'NGRAM_RANGE={NGRAM_RANGE}  |  '
        f'SUBSTRING_MIN_LENGTH={SUBSTRING_MIN_LENGTH}')
cell = ws_s.cell(row=10, column=1, value=note)
cell.font = Font(italic=True, name='Arial', size=9, color='FF666666')
ws_s.merge_cells('A10:C10')

wb.save(OUT_ENRICHED)


# ──────────────────────────────────────────────────────────────────────────────
# 9. FORMATTING — MISMATCH REPORT
# ──────────────────────────────────────────────────────────────────────────────
print("[6/7] Formatting mismatch report ...")

wb2 = load_workbook(OUT_MISMATCH)
ws2 = wb2.active
ws2.title = 'Mismatch Report'
all_cols2   = [c.value for c in ws2[1]]
status_col2 = all_cols2.index('Match_Status') + 1

for col_idx, col_name in enumerate(all_cols2, 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid', fgColor=COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws2.row_dimensions[1].height = 36

for row_idx in range(2, ws2.max_row + 1):
    status_val = ws2.cell(row=row_idx, column=status_col2).value
    fill_color = COLORS['mismatch'] if status_val == 'NAME_MISMATCH' else COLORS['no_match']
    for col_idx in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.fill      = PatternFill('solid', fgColor=fill_color)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

for col_idx in range(1, ws2.max_column + 1):
    vals  = [str(ws2.cell(row=r, column=col_idx).value or '')
             for r in range(1, min(ws2.max_row + 1, 60))]
    width = min(max(max(len(v) for v in vals) + 2, 14), 65)
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.freeze_panes = 'A2'

# ── What remains & why ───────────────────────────────────────────────────────
ws_r = wb2.create_sheet('Remaining Issues Analysis')

analysis = [
    ('Category',         'Description',                                         'Recommended Action'),
    ('Pre-1950 NO_MATCH',
     'Laws from 1920s–1940s simply absent from File2.\nFile2 does not cover this era.',
     'Source from a separate historical legal archive.\nAccept as coverage gap.'),
    ('Duplicate Key Conflict',
     'Same Number+Year maps to a completely different law.\nCosine score typically < 0.15.',
     'Add a 4th key: Law_Type or source category.\nFlag for domain expert validation.'),
    ('Long Descriptive Names',
     'File1 Law_Name contains full legal article text,\nnot a short title.',
     'Regex: extract first noun phrase before first comma\nor parenthesis, then re-run matching.'),
    ('Remaining NAME_MISMATCH\n(score 0.15–0.35)',
     'Partial overlap — possibly same law with\nheavy paraphrasing or modernized wording.',
     'Consider AraBERT semantic embeddings\nas a final recovery pass.'),
]

col_widths3 = [24, 44, 40]
row_colors3 = [COLORS['header'], 'FFFFF2CC', 'FFFCE4D6', 'FFDEEAF1', 'FFE2EFDA']

for r, row in enumerate(analysis, 1):
    for c, val in enumerate(row, 1):
        cell = ws_r.cell(row=r, column=c, value=val)
        cell.fill      = PatternFill('solid', fgColor=row_colors3[r - 1])
        cell.font      = Font(bold=(r == 1 or c == 1),
                               color='FFFFFFFF' if r == 1 else 'FF000000',
                               name='Arial', size=10)
        cell.alignment = Alignment(wrap_text=True, vertical='center',
                                   horizontal='center' if r == 1 else 'left')
        cell.border    = thin_border()
    ws_r.row_dimensions[r].height = 60

for c, w in enumerate(col_widths3, 1):
    ws_r.column_dimensions[get_column_letter(c)].width = w

wb2.save(OUT_MISMATCH)

print("[7/7] Done!\n")
print(f"  → {OUT_ENRICHED}")
print(f"  → {OUT_MISMATCH}")
print(f"\n  Resolved: {resolved:,} / {total:,}  ({100*resolved/total:.1f}%)")
print(f"  Remaining unresolved: {mismatch + no_match:,}")
print(f"\n💡 Tune CONFIG at top of file to adjust thresholds.")
