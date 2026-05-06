"""
Arabic Law Matching — Multi-Strategy Pipeline v5
=================================================
Adds a 5th matching layer specifically for amendment laws
(قانون معدل لقانون X) where File2 stores the original law
with amendments rolled in, not the amendment record itself.

Layer 5 extracts the referenced law name from the amendment title,
then searches File2 for the original law using TF-IDF cosine on
the extracted name — bypassing the Number+Year key mismatch entirely.

Matching layers (applied in order, first match wins):

  Layer 1 — EXACT 3-KEY       : Law_Number + Year + Magazine_Number
  Layer 2 — TF-IDF COSINE     : Law_Number + Year, cosine ≥ threshold
  Layer 3 — SUBSTRING         : Law_Number + Year, name containment
  Layer 4 — LOB LINK          : LobID foreign key → amendments file
  Layer 5 — AMENDMENT REF     : Extract referenced law from title,
                                 search File2 by name (no key constraint)

Outputs:
  enriched_law_v5.xlsx      — File1 enriched, color-coded by layer
  mismatch_report_v5.xlsx   — Truly unresolvable rows

Dependencies:
  pip install pandas openpyxl scikit-learn
"""

import re
import unicodedata
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
FILE1_PATH  = 'law_merged_output_V2.xlsx'
FILE2_PATH  = 'cleaned_v03_merged_updated.xlsx'
FILE3_PATH  = 'amendment_4_5_2026.csv'
OUT_ENRICHED = 'enriched_law_v5.xlsx'
OUT_MISMATCH = 'mismatch_report_v5.xlsx'

COSINE_THRESHOLD    = 0.35   # Layers 2 & 3
AMEND_REF_THRESHOLD = 0.35   # Layer 5 — same as main cosine threshold
SUBSTRING_MIN_LEN   = 6
NGRAM_RANGE         = (2, 4)

ENRICH_MAP_F2 = {
    'Magazine_Number'     : 'magazine_number',
    'Magazine_Date'       : 'magazine_date',
    'Magazine_Page_Number': 'magazine_page',
    'Active_Date'         : 'start_date',
    'Replaced_For'        : 'replaced_for',
}

ENRICH_MAP_F3 = {
    'Magazine_Number'     : 'Magazine_Number',
    'Magazine_Date'       : 'Magazine_Date',
    'Magazine_Page_Number': 'Magazine_Page_Number',
    'Active_Date'         : 'Status_Date',
}

COLORS = {
    'header'  : 'FF1F3864',
    'layer1'  : 'FFC6EFCE',   # green   — exact 3-key
    'layer2'  : 'FFDDEBF7',   # blue    — TF-IDF
    'layer3'  : 'FFEAF3EA',   # mint    — substring
    'layer4'  : 'FFFFD966',   # yellow  — LobID
    'layer5'  : 'FFFFB347',   # orange  — amendment ref
    'mismatch': 'FFFFEB9C',   # amber   — name conflict
    'no_match': 'FFFFC7CE',   # red     — unresolved
    'meta_hdr': 'FF2E75B6',
    'alt'     : 'FFF2F7FF',
}


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────
def normalize_arabic(text: str) -> str:
    if not isinstance(text, str) or not text.strip():
        return ''
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[أإآٱ]', 'ا', text)
    text = re.sub(r'[ئؤ]', 'ء', text)
    text = text.replace('ة', 'ه').replace('ى', 'ي').replace('ـ', '')
    text = re.sub(r'[().,،؛؟!]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def extract_amendment_ref(name: str) -> str:
    """
    From 'قانون معدل لقانون استقلال القضاء' → 'استقلال القضاء'
    From 'قانون معدل للقانون الأساسي'       → 'القانون الأساسي'
    Stops at '/', '(', or end of string.
    Also strips tatweel and collapses whitespace before matching.
    """
    if not isinstance(name, str):
        return ''
    # Clean tatweel and collapse spaces before regex
    name = re.sub(r'ـ+', '', name)
    name = re.sub(r'\s+', ' ', name).strip()
    m = re.search(r'معدل\s+ل(?:قانون|القانون)\s+(.+?)(?:\s*/|\s*\(|$)', name)
    return m.group(1).strip() if m else ''


def to_int_str(val) -> str:
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return ''


def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def enrich_from(record: dict, source_row, enrich_map: dict) -> list:
    enriched = []
    for f1_col, src_col in enrich_map.items():
        val = source_row.get(src_col) if isinstance(source_row, dict) else getattr(source_row, src_col, None)
        if val is None:
            val = source_row[src_col] if src_col in source_row.index else None
        if val is not None and pd.notna(val) and str(val).strip():
            record[f1_col] = val
            enriched.append(f1_col)
    return enriched


# ──────────────────────────────────────────────────────────────────────────────
# 1. LOAD DATA
# ──────────────────────────────────────────────────────────────────────────────
print("[1/7] Loading files ...")
df1 = pd.read_excel(FILE1_PATH)
df2 = pd.read_excel(FILE2_PATH)
df3 = pd.read_csv(FILE3_PATH, encoding='utf-8', low_memory=False)

print(f"      File1 (LOB source)  : {len(df1):,} rows")
print(f"      File2 (enhanced)    : {len(df2):,} rows")
print(f"      File3 (amendments)  : {len(df3):,} rows")

for df, num_col, year_col, mag_col in [
    (df1, 'Law_Number', 'Year', 'Magazine_Number'),
    (df2, 'leg_number', 'year', 'magazine_number'),
]:
    df['_num']  = df[num_col].apply(to_int_str)
    df['_year'] = df[year_col].apply(to_int_str)
    df['_mag']  = df[mag_col].apply(to_int_str)
    df['_key2'] = df['_num'] + '|' + df['_year']
    df['_key3'] = df['_num'] + '|' + df['_year'] + '|' + df['_mag']

df1['_lob']  = df1['LobID'].apply(to_int_str)
df3['_lob']  = df3['LobID'].apply(to_int_str)

df1['_norm'] = df1['Law_Name'].apply(normalize_arabic)
df2['_norm'] = df2['leg_name'].apply(normalize_arabic)

# Extract amendment reference for Layer 5
df1['_amend_ref']      = df1['Law_Name'].apply(extract_amendment_ref)
df1['_amend_ref_norm'] = df1['_amend_ref'].apply(normalize_arabic)


# ──────────────────────────────────────────────────────────────────────────────
# 2. TF-IDF INDICES
# ──────────────────────────────────────────────────────────────────────────────
print("[2/7] Building TF-IDF indices ...")

vectorizer = TfidfVectorizer(
    analyzer    ='char_wb',
    ngram_range = NGRAM_RANGE,
    sublinear_tf= True,
    min_df      = 1,
)
tfidf_f2 = vectorizer.fit_transform(df2['_norm'])

# Lookup structures
df2_by_key2 = df2.groupby('_key2')
df2_by_key3 = df2.set_index('_key3')
df3_by_lob  = df3[df3['_lob'] != ''].groupby('_lob')


# ──────────────────────────────────────────────────────────────────────────────
# 3. MATCHING
# ──────────────────────────────────────────────────────────────────────────────
print("[3/7] Matching (5 layers) ...")

results = []

for _, row1 in df1.iterrows():
    record = row1.to_dict()
    record['_match_layer']     = 'NO_MATCH'
    record['_match_score']     = np.nan
    record['_matched_name']    = ''
    record['_enriched_fields'] = ''

    key2       = row1['_key2']
    key3       = row1['_key3']
    norm1      = row1['_norm']
    lob        = row1['_lob']
    amend_norm = row1['_amend_ref_norm']

    # ── Layer 1: Exact 3-key ─────────────────────────────────────────────────
    if key3 and key3 in df2_by_key3.index:
        best_row = df2_by_key3.loc[[key3]].iloc[0]
        enriched = enrich_from(record, best_row, ENRICH_MAP_F2)
        record.update({'_match_layer': 'EXACT_3KEY', '_match_score': 1.0,
                       '_matched_name': best_row['leg_name'],
                       '_enriched_fields': ', '.join(enriched) or 'none'})
        results.append(record)
        continue

    # ── Layers 2 & 3: File2 candidates by Num+Year ───────────────────────────
    if key2 in df2_by_key2.groups:
        cand_idx   = list(df2_by_key2.groups[key2])
        candidates = df2.loc[cand_idx]
        cand_vecs  = tfidf_f2[cand_idx]

        # Layer 2: TF-IDF cosine
        if norm1:
            q          = vectorizer.transform([norm1])
            cos_scores = cosine_similarity(q, cand_vecs).flatten()
            best_pos   = cos_scores.argmax()
            best_score = cos_scores[best_pos]
            if best_score >= COSINE_THRESHOLD:
                best_row = candidates.iloc[best_pos]
                enriched = enrich_from(record, best_row, ENRICH_MAP_F2)
                record.update({'_match_layer': 'TFIDF',
                               '_match_score': round(float(best_score), 4),
                               '_matched_name': best_row['leg_name'],
                               '_enriched_fields': ', '.join(enriched) or 'none'})
                results.append(record)
                continue

        # Layer 3: Substring
        if norm1 and len(norm1) >= SUBSTRING_MIN_LEN:
            sub_match, sub_score = None, 0.0
            for _, cand_row in candidates.iterrows():
                norm2 = cand_row['_norm']
                if not norm2:
                    continue
                fwd = norm1 in norm2
                rev = norm2 in norm1 and len(norm2) >= SUBSTRING_MIN_LEN
                if fwd or rev:
                    overlap = len(norm1)/len(norm2) if fwd else len(norm2)/len(norm1)
                    if overlap > sub_score:
                        sub_score, sub_match = overlap, cand_row
            if sub_match is not None:
                enriched = enrich_from(record, sub_match, ENRICH_MAP_F2)
                record.update({'_match_layer': 'SUBSTRING',
                               '_match_score': round(sub_score, 4),
                               '_matched_name': sub_match['leg_name'],
                               '_enriched_fields': ', '.join(enriched) or 'none'})
                results.append(record)
                continue

        # Record best cosine for mismatch report
        if norm1:
            q          = vectorizer.transform([norm1])
            cos_scores = cosine_similarity(q, cand_vecs).flatten()
            best_pos   = cos_scores.argmax()
            record['_match_score']  = round(float(cos_scores[best_pos]), 4)
            record['_matched_name'] = candidates.iloc[best_pos]['leg_name']
            record['_match_layer']  = 'NAME_MISMATCH'

    # ── Layer 4: LobID → amendments file ─────────────────────────────────────
    if record['_match_layer'] in ('NO_MATCH', 'NAME_MISMATCH') \
            and lob and lob in df3_by_lob.groups:
        f3_row   = df3_by_lob.get_group(lob).iloc[0]
        enriched = enrich_from(record, f3_row.to_dict(), ENRICH_MAP_F3)
        prior    = record['_match_layer']
        record.update({'_match_layer': f'LOB_LINK (was {prior})',
                       '_matched_name': str(f3_row.get('amendment', '')),
                       '_enriched_fields': ', '.join(enriched) or 'none'})
        results.append(record)
        continue

    # ── Layer 5: Amendment reference name → search all of File2 ──────────────
    # For rows still unresolved that follow the 'قانون معدل لقانون X' pattern.
    # Strategy:
    #   a) Search all of File2 by extracted reference name using TF-IDF cosine.
    #   b) Among top-20 candidates, filter to rows where year <= amendment year
    #      (an amendment can only amend a law that already exists).
    #   c) Pick the highest-scoring eligible candidate.
    if record['_match_layer'] in ('NO_MATCH', 'NAME_MISMATCH') \
            and amend_norm and len(amend_norm) >= SUBSTRING_MIN_LEN:

        q          = vectorizer.transform([amend_norm])
        cos_scores = cosine_similarity(q, tfidf_f2).flatten()

        # Get top 20 candidates
        top20_idx  = cos_scores.argsort()[-20:][::-1]
        top20      = df2.iloc[top20_idx].copy()
        top20['_cs'] = cos_scores[top20_idx]

        # Tiebreaker: original law must predate or equal the amendment year
        f1_year = row1['_year']
        eligible = top20[
            top20['_year'].apply(
                lambda y: y != '' and f1_year != '' and int(y) <= int(f1_year)
            )
        ]

        candidates_to_check = eligible if len(eligible) > 0 else top20

        best_row   = candidates_to_check.iloc[0]
        best_score = best_row['_cs']

        if best_score >= AMEND_REF_THRESHOLD:
            enriched = enrich_from(record, best_row, ENRICH_MAP_F2)
            prior    = record['_match_layer']
            record.update({'_match_layer': f'AMEND_REF (was {prior})',
                           '_match_score': round(float(best_score), 4),
                           '_matched_name': best_row['leg_name'],
                           '_enriched_fields': ', '.join(enriched) or 'none'})

    results.append(record)


result_df = pd.DataFrame(results)


# ──────────────────────────────────────────────────────────────────────────────
# 4. SUMMARY
# ──────────────────────────────────────────────────────────────────────────────
def cnt(df, prefix):
    return df['_match_layer'].str.startswith(prefix).sum()

total    = len(result_df)
layer1   = cnt(result_df, 'EXACT_3KEY')
layer2   = cnt(result_df, 'TFIDF')
layer3   = cnt(result_df, 'SUBSTRING')
layer4   = cnt(result_df, 'LOB_LINK')
layer5   = cnt(result_df, 'AMEND_REF')
mismatch = cnt(result_df, 'NAME_MISMATCH')
no_match = cnt(result_df, 'NO_MATCH')
resolved = layer1 + layer2 + layer3 + layer4 + layer5

print(f"\n{'='*62}")
print(f"  MATCHING SUMMARY  (cosine={COSINE_THRESHOLD}, amend_ref={AMEND_REF_THRESHOLD})")
print(f"{'='*62}")
print(f"  Total File1 records            :  {total:>6,}")
print(f"  ─────────────────────────────────────────────")
print(f"  ✅ Layer 1 — Exact 3-key       :  {layer1:>6,}  ({100*layer1/total:.1f}%)")
print(f"  ✅ Layer 2 — TF-IDF cosine     :  {layer2:>6,}  ({100*layer2/total:.1f}%)")
print(f"  ✅ Layer 3 — Substring         :  {layer3:>6,}  ({100*layer3/total:.1f}%)")
print(f"  ✅ Layer 4 — LobID Amendment   :  {layer4:>6,}  ({100*layer4/total:.1f}%)")
print(f"  ✅ Layer 5 — Amendment Ref     :  {layer5:>6,}  ({100*layer5/total:.1f}%)")
print(f"  ─────────────────────────────────────────────")
print(f"  ✅ TOTAL RESOLVED              :  {resolved:>6,}  ({100*resolved/total:.1f}%)")
print(f"  ⚠️  Name mismatch               :  {mismatch:>6,}  ({100*mismatch/total:.1f}%)")
print(f"  ❌ No match (truly absent)     :  {no_match:>6,}  ({100*no_match/total:.1f}%)")
print(f"{'='*62}\n")


# ──────────────────────────────────────────────────────────────────────────────
# 5. EXPORT RAW
# ──────────────────────────────────────────────────────────────────────────────
print("[4/7] Exporting raw Excel files ...")

output_cols = [c for c in df1.columns if not c.startswith('_')]
meta_cols   = ['_match_layer', '_match_score', '_matched_name', '_enriched_fields']

result_df[output_cols + meta_cols].to_excel(OUT_ENRICHED, index=False)

mismatch_df = result_df[
    result_df['_match_layer'].isin(['NAME_MISMATCH', 'NO_MATCH'])
][['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
   '_match_layer', '_match_score', '_matched_name']].copy()
mismatch_df.columns = ['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
                        'Match_Status', 'Best_Score', 'Best_Candidate']
mismatch_df.to_excel(OUT_MISMATCH, index=False)


# ──────────────────────────────────────────────────────────────────────────────
# 6. FORMAT — ENRICHED FILE
# ──────────────────────────────────────────────────────────────────────────────
print("[5/7] Formatting enriched file ...")

def layer_color(layer_val):
    if not isinstance(layer_val, str): return COLORS['alt']
    if layer_val.startswith('EXACT_3KEY') : return COLORS['layer1']
    if layer_val.startswith('TFIDF')      : return COLORS['layer2']
    if layer_val.startswith('SUBSTRING')  : return COLORS['layer3']
    if layer_val.startswith('LOB_LINK')   : return COLORS['layer4']
    if layer_val.startswith('AMEND_REF')  : return COLORS['layer5']
    if layer_val.startswith('NAME_MISMATCH'): return COLORS['mismatch']
    return COLORS['no_match']

wb = load_workbook(OUT_ENRICHED)
ws = wb.active
ws.title = 'Enriched Data'

all_cols       = [c.value for c in ws[1]]
meta_col_set   = set(meta_cols)
status_col_idx = all_cols.index('_match_layer') + 1

for col_idx, col_name in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid', fgColor=COLORS['meta_hdr']
                                 if col_name in meta_col_set else COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[1].height = 36

for row_idx in range(2, ws.max_row + 1):
    layer = ws.cell(row=row_idx, column=status_col_idx).value
    fc    = layer_color(layer)
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill      = PatternFill('solid', fgColor=fc)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

for col_idx, col_name in enumerate(all_cols, 1):
    vals  = [str(ws.cell(row=r, column=col_idx).value or '')
             for r in range(1, min(ws.max_row + 1, 60))]
    ws.column_dimensions[get_column_letter(col_idx)].width = min(
        max(max(len(v) for v in vals) + 2, 12), 50)

ws.freeze_panes = 'A2'

# Summary sheet
ws_s = wb.create_sheet('Summary')
summary_rows = [
    ('Layer',                             'Count',   'Share',                       COLORS['header']),
    ('Total File1 Records',                total,    '100%',                         COLORS['header']),
    ('Layer 1 — Exact 3-Key',              layer1,  f'{100*layer1/total:.1f}%',     COLORS['layer1']),
    ('Layer 2 — TF-IDF Cosine',            layer2,  f'{100*layer2/total:.1f}%',     COLORS['layer2']),
    ('Layer 3 — Substring',                layer3,  f'{100*layer3/total:.1f}%',     COLORS['layer3']),
    ('Layer 4 — LobID Amendment Link',     layer4,  f'{100*layer4/total:.1f}%',     COLORS['layer4']),
    ('Layer 5 — Amendment Ref Name',       layer5,  f'{100*layer5/total:.1f}%',     COLORS['layer5']),
    ('TOTAL RESOLVED',                    resolved, f'{100*resolved/total:.1f}%',   'FFA9D18C'),
    ('Name Mismatch',                     mismatch, f'{100*mismatch/total:.1f}%',   COLORS['mismatch']),
    ('No Match (truly absent)',            no_match, f'{100*no_match/total:.1f}%',  COLORS['no_match']),
]

for r, (label, count, share, color) in enumerate(summary_rows, 1):
    for c, val in enumerate([label, count, share], 1):
        cell = ws_s.cell(row=r, column=c, value=val)
        cell.fill      = PatternFill('solid', fgColor=color)
        cell.font      = Font(bold=(r in (1, 8)),
                               color='FFFFFFFF' if r <= 2 else 'FF000000',
                               name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border()
    ws_s.row_dimensions[r].height = 26

ws_s.column_dimensions['A'].width = 36
ws_s.column_dimensions['B'].width = 12
ws_s.column_dimensions['C'].width = 12

# Legend
ws_s.cell(row=12, column=1, value='Layer 5 note: matches amendment laws (قانون معدل لقانون X) '
          'by searching File2 for the referenced original law name, bypassing key mismatch.')
ws_s.cell(row=12, column=1).font = Font(italic=True, name='Arial', size=9, color='FF666666')
ws_s.merge_cells('A12:C12')
ws_s.row_dimensions[12].height = 30

wb.save(OUT_ENRICHED)


# ──────────────────────────────────────────────────────────────────────────────
# 7. FORMAT — MISMATCH REPORT
# ──────────────────────────────────────────────────────────────────────────────
print("[6/7] Formatting mismatch report ...")

wb2 = load_workbook(OUT_MISMATCH)
ws2 = wb2.active
ws2.title = 'Mismatch Report'
all_cols2   = [c.value for c in ws2[1]]
status_col2 = all_cols2.index('Match_Status') + 1

for col_idx in range(1, ws2.max_column + 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid', fgColor=COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws2.row_dimensions[1].height = 36

for row_idx in range(2, ws2.max_row + 1):
    status_val = ws2.cell(row=row_idx, column=status_col2).value
    fc = COLORS['mismatch'] if status_val == 'NAME_MISMATCH' else COLORS['no_match']
    for col_idx in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.fill      = PatternFill('solid', fgColor=fc)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

for col_idx in range(1, ws2.max_column + 1):
    vals  = [str(ws2.cell(row=r, column=col_idx).value or '')
             for r in range(1, min(ws2.max_row + 1, 60))]
    ws2.column_dimensions[get_column_letter(col_idx)].width = min(
        max(max(len(v) for v in vals) + 2, 14), 65)
ws2.freeze_panes = 'A2'

wb2.save(OUT_MISMATCH)

print("[7/7] Done!\n")
print(f"  → {OUT_ENRICHED}")
print(f"  → {OUT_MISMATCH}")
print(f"\n  Resolved  : {resolved:,} / {total:,}  ({100*resolved/total:.1f}%)")
print(f"  Remaining : {mismatch + no_match:,}")
print(f"\n  Layer 5 recovered: {layer5:,} amendment-reference rows")