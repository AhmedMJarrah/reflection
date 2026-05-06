"""
Arabic Law Name Matching — TF-IDF Character N-gram + Cosine Similarity
=======================================================================
Replaces fuzzy (Levenshtein) matching with:
  1. Arabic text normalization  (Alef/Hamza/Tashkeel)
  2. Character trigram TF-IDF vectorization
  3. Cosine similarity for name matching

Pipeline:
  - Exact match first  →  Law_Number + Year
  - Among candidates   →  pick best cosine similarity score
  - Threshold          →  COSINE_THRESHOLD (default 0.35, tune as needed)
  - Below threshold    →  flagged as NAME_MISMATCH for review

Outputs (code only — no Excel generated):
  enriched_law_output.xlsx    →  File1 enriched with File2 data
  mismatch_report.xlsx        →  Unresolved rows with scores + best candidate

Dependencies:
  pip install pandas openpyxl scikit-learn
  (no GPU, no large models, no paid APIs)
"""

import re
import unicodedata
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
FILE1_PATH       = r'C:\Users\user\Desktop\ref\law_merged_output_V2.xlsx'       # LOB source
FILE2_PATH       = r'C:\Users\user\Desktop\ref\cleaned_v03_merged_updated.xlsx'  # Enhanced
OUT_ENRICHED     = 'enriched_law_tfidf.xlsx'
OUT_MISMATCH     = 'mismatch_report_tfidf.xlsx'

COSINE_THRESHOLD = 0.35   # tune: lower = more matches, higher = more precise
NGRAM_RANGE      = (2, 4) # character n-gram range (bigrams to 4-grams)

# Enrichment column map  {File1_col: File2_col}
ENRICH_MAP = {
    'Magazine_Number'     : 'magazine_number',
    'Magazine_Date'       : 'magazine_date',
    'Magazine_Page_Number': 'magazine_page',
    'Active_Date'         : 'start_date',
    'Replaced_For'        : 'replaced_for',
}

COLORS = {
    'header'   : '1F3864',
    'matched'  : 'C6EFCE',
    'mismatch' : 'FFEB9C',
    'no_match' : 'FFC7CE',
    'meta_hdr' : '2E75B6',
    'alt'      : 'F2F7FF',
}


# ──────────────────────────────────────────────────────────────────────────────
# 1. ARABIC TEXT NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
def normalize_arabic(text: str) -> str:
    """
    Normalize Arabic text for robust string comparison:
      - Strip tashkeel (diacritics / harakat)
      - Unify Alef variants  (أ إ آ ا → ا)
      - Unify Hamza variants (ئ ؤ → ء)
      - Unify Teh marbuta    (ة → ه)
      - Unify Yeh variants   (ى → ي)
      - Remove tatweel       (ـ)
      - Collapse whitespace
    """
    if not isinstance(text, str) or not text.strip():
        return ''

    # Remove tashkeel (Unicode category Mn = non-spacing marks)
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')

    # Alef variants → bare Alef
    text = re.sub(r'[أإآٱ]', 'ا', text)
    # Hamza variants
    text = re.sub(r'[ئؤ]', 'ء', text)
    # Teh marbuta → Heh
    text = text.replace('ة', 'ه')
    # Dotless Yeh → Yeh
    text = text.replace('ى', 'ي')
    # Tatweel
    text = text.replace('ـ', '')
    # Collapse whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    return text


# ──────────────────────────────────────────────────────────────────────────────
# 2. HELPER: NUMERIC KEY NORMALIZATION
# ──────────────────────────────────────────────────────────────────────────────
def to_int_str(val) -> str:
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return ''


def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


# ──────────────────────────────────────────────────────────────────────────────
# 3. LOAD & PREPARE DATA
# ──────────────────────────────────────────────────────────────────────────────
print("[1/6] Loading files ...")
df1 = pd.read_excel(FILE1_PATH)
df2 = pd.read_excel(FILE2_PATH)
print(f"      File1: {len(df1):,} rows | File2: {len(df2):,} rows")

# Numeric keys
for df, num_col, year_col in [(df1, 'Law_Number', 'Year'),
                               (df2, 'leg_number', 'year')]:
    df['_num']  = df[num_col].apply(to_int_str)
    df['_year'] = df[year_col].apply(to_int_str)
    df['_key']  = df['_num'] + '|' + df['_year']

# Normalized name for TF-IDF
df1['_norm_name'] = df1['Law_Name'].apply(normalize_arabic)
df2['_norm_name'] = df2['leg_name'].apply(normalize_arabic)


# ──────────────────────────────────────────────────────────────────────────────
# 4. BUILD TF-IDF VECTORIZER ON ALL FILE2 NAMES
# ──────────────────────────────────────────────────────────────────────────────
print("[2/6] Building TF-IDF character n-gram index on File2 names ...")

vectorizer = TfidfVectorizer(
    analyzer   = 'char_wb',   # character n-grams with word boundaries
    ngram_range= NGRAM_RANGE,
    sublinear_tf= True,       # log(tf) — dampens effect of very frequent tokens
    min_df     = 1,
)

# Fit on all File2 normalized names
tfidf_matrix_f2 = vectorizer.fit_transform(df2['_norm_name'])


# ──────────────────────────────────────────────────────────────────────────────
# 5. MATCH + ENRICH
# ──────────────────────────────────────────────────────────────────────────────
print("[3/6] Matching and enriching ...")

df2_grouped = df2.groupby('_key')

results = []

for idx, row1 in df1.iterrows():
    key    = row1['_key']
    record = row1.to_dict()
    record['_match_status']      = 'NO_MATCH'
    record['_cosine_similarity'] = np.nan
    record['_matched_leg_name']  = ''
    record['_enriched_fields']   = ''

    if key not in df2_grouped.groups:
        results.append(record)
        continue

    # Get candidate rows from File2 with same Number+Year
    candidate_indices = df2_grouped.groups[key]
    candidates        = df2.loc[candidate_indices]

    # TF-IDF cosine similarity: query = File1 name
    query_vec    = vectorizer.transform([row1['_norm_name']])
    cand_vecs    = tfidf_matrix_f2[candidate_indices]
    cos_scores   = cosine_similarity(query_vec, cand_vecs).flatten()

    best_pos     = cos_scores.argmax()
    best_score   = cos_scores[best_pos]
    best_row     = candidates.iloc[best_pos]

    record['_cosine_similarity'] = round(float(best_score), 4)
    record['_matched_leg_name']  = best_row['leg_name']

    if best_score >= COSINE_THRESHOLD:
        record['_match_status'] = 'MATCHED'
        enriched = []
        for f1_col, f2_col in ENRICH_MAP.items():
            f2_val = best_row[f2_col]
            if pd.notna(f2_val) and f2_val != '':
                record[f1_col] = f2_val
                enriched.append(f1_col)
        record['_enriched_fields'] = ', '.join(enriched) if enriched else 'none'
    else:
        record['_match_status'] = 'NAME_MISMATCH'

    results.append(record)

result_df = pd.DataFrame(results)


# ──────────────────────────────────────────────────────────────────────────────
# 6. SUMMARY
# ──────────────────────────────────────────────────────────────────────────────
total    = len(result_df)
matched  = (result_df['_match_status'] == 'MATCHED').sum()
mismatch = (result_df['_match_status'] == 'NAME_MISMATCH').sum()
no_match = (result_df['_match_status'] == 'NO_MATCH').sum()

print(f"\n{'='*56}")
print(f"  MATCHING SUMMARY  (threshold = {COSINE_THRESHOLD})")
print(f"{'='*56}")
print(f"  Total File1 records  :  {total:>6,}")
print(f"  ✅ Matched & enriched :  {matched:>6,}  ({100*matched/total:.1f}%)")
print(f"  ⚠️  Name mismatch      :  {mismatch:>6,}  ({100*mismatch/total:.1f}%)")
print(f"  ❌ No match found     :  {no_match:>6,}  ({100*no_match/total:.1f}%)")
print(f"{'='*56}\n")


# ──────────────────────────────────────────────────────────────────────────────
# 7. EXPORT RAW FILES
# ──────────────────────────────────────────────────────────────────────────────
print("[4/6] Exporting raw Excel files ...")

output_cols = [c for c in df1.columns if not c.startswith('_')]
meta_cols   = ['_match_status', '_cosine_similarity',
               '_matched_leg_name', '_enriched_fields']

enriched_export = result_df[output_cols + meta_cols].copy()
enriched_export.to_excel(OUT_ENRICHED, index=False)

mismatch_df = result_df[
    result_df['_match_status'].isin(['NAME_MISMATCH', 'NO_MATCH'])
][['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
   '_match_status', '_cosine_similarity', '_matched_leg_name']].copy()

mismatch_df.columns = ['Law_Name', 'Law_Number', 'Year', 'Magazine_Number',
                        'Match_Status', 'Cosine_Similarity', 'Best_Candidate_in_File2']

mismatch_df.to_excel(OUT_MISMATCH, index=False)


# ──────────────────────────────────────────────────────────────────────────────
# 8. PROFESSIONAL FORMATTING — ENRICHED FILE
# ──────────────────────────────────────────────────────────────────────────────
print("[5/6] Applying formatting to enriched file ...")

wb = load_workbook(OUT_ENRICHED)
ws = wb.active
ws.title = 'Enriched Data'

all_cols       = [cell.value for cell in ws[1]]
meta_col_set   = set(meta_cols)
status_col_idx = all_cols.index('_match_status') + 1

# Header
for col_idx, col_name in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid',
                                  fgColor=COLORS['meta_hdr'] if col_name in meta_col_set
                                  else COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws.row_dimensions[1].height = 36

# Data rows
STATUS_COLOR = {
    'MATCHED'       : COLORS['matched'],
    'NAME_MISMATCH' : COLORS['mismatch'],
    'NO_MATCH'      : COLORS['no_match'],
}

for row_idx in range(2, ws.max_row + 1):
    status_val  = ws.cell(row=row_idx, column=status_col_idx).value
    fill_color  = STATUS_COLOR.get(status_val, COLORS['alt'] if row_idx % 2 == 0 else None)

    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if fill_color:
            cell.fill = PatternFill('solid', fgColor=fill_color)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

# Column widths
for col_idx, col_name in enumerate(all_cols, 1):
    sample_vals = [str(ws.cell(row=r, column=col_idx).value or '')
                   for r in range(1, min(ws.max_row + 1, 60))]
    width = min(max(max(len(v) for v in sample_vals) + 2, 12), 50)
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.freeze_panes = 'A2'

# Summary sheet
ws_sum = wb.create_sheet('Summary')
summary_rows = [
    ('Metric',                       'Count',   'Share'),
    ('Total File1 Records',           total,    '100%'),
    ('✅ Matched & Enriched',          matched,  f'{100*matched/total:.1f}%'),
    ('⚠️  Name Mismatch (low cosine)', mismatch, f'{100*mismatch/total:.1f}%'),
    ('❌ No Match (key not found)',    no_match, f'{100*no_match/total:.1f}%'),
]
row_fills = [COLORS['header'], None,
             COLORS['matched'], COLORS['mismatch'], COLORS['no_match']]

for r, (row, fill) in enumerate(zip(summary_rows, row_fills), 1):
    for c, val in enumerate(row, 1):
        cell = ws_sum.cell(row=r, column=c, value=val)
        cell.font      = Font(bold=(r == 1), color='FFFFFF' if r == 1 else '000000',
                               name='Arial', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border()
        if fill:
            cell.fill = PatternFill('solid', fgColor=fill)
    ws_sum.row_dimensions[r].height = 26

ws_sum.column_dimensions['A'].width = 36
ws_sum.column_dimensions['B'].width = 12
ws_sum.column_dimensions['C'].width = 12

# Config note
ws_sum.cell(row=7, column=1,
            value=f'Cosine threshold used: {COSINE_THRESHOLD}  |  N-gram range: {NGRAM_RANGE}  |  Vectorizer: char_wb TF-IDF')
ws_sum.cell(row=7, column=1).font = Font(italic=True, name='Arial', size=10, color='666666')
ws_sum.merge_cells('A7:C7')

wb.save(OUT_ENRICHED)


# ──────────────────────────────────────────────────────────────────────────────
# 9. PROFESSIONAL FORMATTING — MISMATCH REPORT
# ──────────────────────────────────────────────────────────────────────────────
print("[6/6] Applying formatting to mismatch report ...")

wb2 = load_workbook(OUT_MISMATCH)
ws2 = wb2.active
ws2.title = 'Mismatch Report'

all_cols2      = [cell.value for cell in ws2[1]]
status_col2    = all_cols2.index('Match_Status') + 1

for col_idx, col_name in enumerate(all_cols2, 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.fill      = PatternFill('solid', fgColor=COLORS['header'])
    cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = thin_border()
ws2.row_dimensions[1].height = 36

for row_idx in range(2, ws2.max_row + 1):
    status_val = ws2.cell(row=row_idx, column=status_col2).value
    fill_color = COLORS['mismatch'] if status_val == 'NAME_MISMATCH' else COLORS['no_match']
    for col_idx in range(1, ws2.max_column + 1):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.fill      = PatternFill('solid', fgColor=fill_color)
        cell.font      = Font(name='Arial', size=9)
        cell.alignment = Alignment(vertical='center')
        cell.border    = thin_border()

for col_idx, col_name in enumerate(all_cols2, 1):
    sample_vals = [str(ws2.cell(row=r, column=col_idx).value or '')
                   for r in range(1, min(ws2.max_row + 1, 60))]
    width = min(max(max(len(v) for v in sample_vals) + 2, 14), 65)
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

ws2.freeze_panes = 'A2'

# Techniques sheet
ws_rec = wb2.create_sheet('Techniques & Next Steps')
tech_rows = [
    ('Priority', 'Technique', 'Description', 'Expected Recovery'),
    ('✅ Applied',
     'TF-IDF char n-gram\n+ Cosine Similarity',
     'Character (2-4)-gram TF-IDF with Arabic normalization.\nRobust to word order, extra words, Hamza/Alef variants.',
     'Primary method — ~60-70% of cases'),
    ('🔧 Next',
     'Partial containment check',
     'Check if File1 short name is a substring of\nFile2 long formal title after normalization.',
     'Catches ~15% more (formal title wrapping)'),
    ('🤖 Advanced',
     'AraBERT / CAMeL-BERT\nEmbeddings + Cosine',
     'Semantic vector similarity. Handles paraphrased\nor modernized law titles.',
     '~10% additional for hard cases'),
    ('👁️ Final',
     'Manual Review',
     'Export remaining rows sorted by cosine score.\nDomain expert validates edge cases.',
     'Final QA layer — <5% of total'),
]
col_widths2 = [14, 28, 42, 26]
row_fills2  = [COLORS['header'], COLORS['matched'],
               COLORS['mismatch'], COLORS['no_match'], 'EDE7F6']

for r, (row, fill) in enumerate(zip(tech_rows, row_fills2), 1):
    for c, val in enumerate(row, 1):
        cell = ws_rec.cell(row=r, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical='center',
                                   horizontal='center' if r == 1 else 'left')
        cell.border    = thin_border()
        cell.font      = Font(bold=(r == 1 or c == 2), name='Arial', size=10,
                               color='FFFFFF' if r == 1 else '000000')
        cell.fill      = PatternFill('solid', fgColor=fill)
    ws_rec.row_dimensions[r].height = 52

for c, w in enumerate(col_widths2, 1):
    ws_rec.column_dimensions[get_column_letter(c)].width = w

wb2.save(OUT_MISMATCH)

print("\n✅ Done!")
print(f"   → {OUT_ENRICHED}")
print(f"   → {OUT_MISMATCH}")
print(f"\n💡 Tip: Adjust COSINE_THRESHOLD in CONFIG section to tune precision vs. recall.")
print(f"        Current: {COSINE_THRESHOLD}  |  Try 0.25 to recover more matches.")
