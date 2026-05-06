"""
Phase 1 — Data Exploration & Understanding
===========================================
Explores both files deeply and produces a clear picture of:
  1. Basic statistics
  2. Column completeness
  3. Duplicate analysis (within each file)
  4. Cross-file overlap analysis
  5. The extra rows in File2
  6. The end_date column

Output: phase1_exploration_report.xlsx
  - Sheet 1: Overview
  - Sheet 2: File1 Duplicates
  - Sheet 3: File2 Duplicates
  - Sheet 4: Cross-File Analysis
  - Sheet 5: Extra Laws in File2
  - Sheet 6: end_date Analysis

Run: python phase1_exploration.py
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ──────────────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────────────
FILE1_PATH = 'law_merged_output_V2.xlsx'
FILE2_PATH = 'cleaned_v03_merged_updated.xlsx'
OUT_FILE   = 'phase1_exploration_report.xlsx'

COLORS = {
    'header'   : 'FF1F3864',
    'subhdr'   : 'FF2E75B6',
    'green'    : 'FFC6EFCE',
    'amber'    : 'FFFFEB9C',
    'red'      : 'FFFFC7CE',
    'blue'     : 'FFDDEBF7',
    'purple'   : 'FFEAD1DC',
    'alt'      : 'FFF2F7FF',
    'white'    : 'FFFFFFFF',
}

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def fmt(cell, fill=None, bold=False, color='FF000000', size=10,
        align='left', wrap=False):
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    cell.font      = Font(bold=bold, color=color, name='Arial', size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap)
    cell.border    = thin_border()

def to_int_str(v):
    try: return str(int(float(v)))
    except: return ''

# ──────────────────────────────────────────────────────────────────────────────
# LOAD
# ──────────────────────────────────────────────────────────────────────────────
print("[1/6] Loading files ...")
df1 = pd.read_excel(FILE1_PATH)
df2 = pd.read_excel(FILE2_PATH)
print(f"      File1: {len(df1):,} rows  x  {len(df1.columns)} columns")
print(f"      File2: {len(df2):,} rows  x  {len(df2.columns)} columns")

# Normalise keys
df1['_num']  = df1['Law_Number'].apply(to_int_str)
df1['_year'] = df1['Year'].apply(to_int_str)
df1['_name'] = df1['Law_Name'].astype(str).str.strip()
df1['_key3'] = df1['_name'] + '|' + df1['_num'] + '|' + df1['_year']
df1['_key2'] = df1['_num'] + '|' + df1['_year']

df2['_num']  = df2['leg_number'].apply(to_int_str)
df2['_year'] = df2['year'].apply(to_int_str)
df2['_name'] = df2['leg_name'].astype(str).str.strip()
df2['_key3'] = df2['_name'] + '|' + df2['_num'] + '|' + df2['_year']
df2['_key2'] = df2['_num'] + '|' + df2['_year']

# ──────────────────────────────────────────────────────────────────────────────
# ANALYSIS
# ──────────────────────────────────────────────────────────────────────────────
print("[2/6] Analysing duplicates ...")

# File1 duplicates
f1_full_dup  = df1[df1.duplicated(keep=False)]                                  # 100% identical
f1_key3_dup  = df1[df1.duplicated(subset=['_name','_num','_year'], keep=False)]  # same name+num+year
f1_key2_dup  = df1[df1.duplicated(subset=['_num','_year'], keep=False)]          # same num+year only

# File2 duplicates
f2_full_dup  = df2[df2.duplicated(keep=False)]
f2_key3_dup  = df2[df2.duplicated(subset=['_name','_num','_year'], keep=False)]
f2_key2_dup  = df2[df2.duplicated(subset=['_num','_year'], keep=False)]

# Cross-file overlap
f1_keys3_in_f2 = df1['_key3'].isin(df2['_key3'])   # exact name+num+year match
f1_keys2_in_f2 = df1['_key2'].isin(df2['_key2'])   # num+year match (may differ in name)
f2_keys3_in_f1 = df2['_key3'].isin(df1['_key3'])
f2_keys2_in_f1 = df2['_key2'].isin(df1['_key2'])

# Extra in File2 (not in File1 by exact name+num+year)
extra_f2 = df2[~f2_keys3_in_f1].copy()

# File1 key3 duplicate groups — detail
f1_dup_groups = []
for key, grp in df1[df1.duplicated(subset=['_name','_num','_year'], keep=False)]\
        .groupby(['_name','_num','_year']):
    enrich_cols = ['Magazine_Number','Magazine_Date','Magazine_Page_Number',
                   'Active_Date','Replaced_For','Status','LobID','IsPublished']
    diff_cols = [c for c in enrich_cols if c in grp.columns
                 and grp[c].fillna('NULL').astype(str).nunique() > 1]
    null_counts = grp.isnull().sum(axis=1).tolist()
    f1_dup_groups.append({
        'Law_Name'         : key[0][:80],
        'Law_Number'       : key[1],
        'Year'             : key[2],
        'Occurrences'      : len(grp),
        'pmk_IDs'          : ', '.join(grp['pmk_ID'].astype(str).tolist()),
        'Columns_Differ'   : ', '.join(diff_cols) if diff_cols else 'None — fully identical',
        'Null_Count_per_Row': ' | '.join(str(n) for n in null_counts),
        'Recommendation'   : 'MERGE (keep most complete row, fill nulls from others)'
                              if diff_cols else 'DROP DUPLICATE (keep one)',
    })

f1_dup_df = pd.DataFrame(f1_dup_groups)

# File2 key3 duplicate groups — detail
f2_dup_groups = []
for key, grp in df2[df2.duplicated(subset=['_name','_num','_year'], keep=False)]\
        .groupby(['_name','_num','_year']):
    enrich_cols2 = ['magazine_number','magazine_date','magazine_page',
                    'start_date','end_date','replaced_for','cancelled_by','status']
    diff_cols = [c for c in enrich_cols2 if c in grp.columns
                 and grp[c].fillna('NULL').astype(str).nunique() > 1]
    null_counts = grp.isnull().sum(axis=1).tolist()
    f2_dup_groups.append({
        'leg_name'         : key[0][:80],
        'leg_number'       : key[1],
        'year'             : key[2],
        'Occurrences'      : len(grp),
        'Columns_Differ'   : ', '.join(diff_cols) if diff_cols else 'None — fully identical',
        'Null_Count_per_Row': ' | '.join(str(n) for n in null_counts),
        'Recommendation'   : 'MERGE (keep most complete row, fill nulls from others)'
                              if diff_cols else 'DROP DUPLICATE (keep one)',
    })

f2_dup_df = pd.DataFrame(f2_dup_groups)

# Cross-file analysis
cross_data = {
    'Description'                                          : 'Count',
    'File1 rows with EXACT match in File2 (Name+Num+Year)': int(f1_keys3_in_f2.sum()),
    'File1 rows with Num+Year match in File2 (any name)'  : int(f1_keys2_in_f2.sum()),
    'File1 rows with NO match in File2 at all'            : int((~f1_keys2_in_f2).sum()),
    'File2 rows with EXACT match in File1 (Name+Num+Year)': int(f2_keys3_in_f1.sum()),
    'File2 rows with Num+Year match in File1 (any name)'  : int(f2_keys2_in_f1.sum()),
    'File2 rows with NO match in File1 at all'            : int((~f2_keys2_in_f1).sum()),
    'Truly new laws in File2 (no Name+Num+Year in File1)' : len(extra_f2),
}

# Extra laws detail
extra_f2_display = extra_f2[['leg_name','leg_number','year',
                               'magazine_number','start_date','end_date','status']].copy()
extra_f2_display.columns = ['Law_Name','Law_Number','Year',
                             'Magazine_Number','Start_Date','End_Date','Status']

# end_date analysis
end_existing = df2[f2_keys3_in_f1]['end_date']  # end_date for laws already in File1
end_extra    = df2[~f2_keys3_in_f1]['end_date']  # end_date for new laws

end_analysis = {
    'end_date non-null in ALL of File2'               : int(df2['end_date'].notna().sum()),
    'end_date non-null for laws ALREADY in File1'     : int(end_existing.notna().sum()),
    'end_date non-null for NEW laws (not in File1)'   : int(end_extra.notna().sum()),
    'end_date value "0" (likely placeholder for null)': int((df2['end_date'] == '0').sum()),
    'Unique end_date values sample'                   : ', '.join(
        df2['end_date'].dropna().unique()[:8].tolist()),
}

print("[3/6] Building column completeness stats ...")

# Column completeness
f1_completeness = pd.DataFrame({
    'Column'      : df1.columns[~df1.columns.str.startswith('_')],
    'Non_Null'    : [df1[c].notna().sum() for c in df1.columns if not c.startswith('_')],
    'Null'        : [df1[c].isna().sum()  for c in df1.columns if not c.startswith('_')],
    'Completeness': [f"{100*df1[c].notna().sum()/len(df1):.1f}%"
                     for c in df1.columns if not c.startswith('_')],
})

f2_completeness = pd.DataFrame({
    'Column'      : df2.columns[~df2.columns.str.startswith('_')],
    'Non_Null'    : [df2[c].notna().sum() for c in df2.columns if not c.startswith('_')],
    'Null'        : [df2[c].isna().sum()  for c in df2.columns if not c.startswith('_')],
    'Completeness': [f"{100*df2[c].notna().sum()/len(df2):.1f}%"
                     for c in df2.columns if not c.startswith('_')],
})

# ──────────────────────────────────────────────────────────────────────────────
# EXPORT
# ──────────────────────────────────────────────────────────────────────────────
print("[4/6] Writing Excel report ...")

with pd.ExcelWriter(OUT_FILE, engine='openpyxl') as writer:

    # Sheet 1 — Overview (we'll build this manually after)
    pd.DataFrame().to_excel(writer, sheet_name='Overview', index=False)

    # Sheet 2 — File1 column completeness
    f1_completeness.to_excel(writer, sheet_name='File1 Completeness', index=False)

    # Sheet 3 — File2 column completeness
    f2_completeness.to_excel(writer, sheet_name='File2 Completeness', index=False)

    # Sheet 4 — File1 duplicates
    f1_dup_df.to_excel(writer, sheet_name='File1 Duplicates', index=False)

    # Sheet 5 — File2 duplicates
    f2_dup_df.to_excel(writer, sheet_name='File2 Duplicates', index=False)

    # Sheet 6 — Cross-file analysis
    pd.DataFrame(list(cross_data.items()),
                 columns=['Description','Count'])\
      .to_excel(writer, sheet_name='Cross-File Analysis', index=False)

    # Sheet 7 — Extra laws in File2
    extra_f2_display.to_excel(writer, sheet_name='Extra Laws in File2', index=False)

    # Sheet 8 — end_date analysis
    pd.DataFrame(list(end_analysis.items()),
                 columns=['Metric','Value'])\
      .to_excel(writer, sheet_name='end_date Analysis', index=False)

# ──────────────────────────────────────────────────────────────────────────────
# FORMAT
# ──────────────────────────────────────────────────────────────────────────────
print("[5/6] Formatting ...")

wb = load_workbook(OUT_FILE)

def style_sheet(ws, header_fill=None, row_colors=None, col_widths=None):
    """Apply standard formatting to a sheet."""
    hf = header_fill or COLORS['header']
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        fmt(cell, fill=hf, bold=True, color='FFFFFFFF', align='center', size=10)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            fill = COLORS['alt'] if row_idx % 2 == 0 else COLORS['white']
            if row_colors:
                fill = row_colors.get(row_idx, fill)
            fmt(cell, fill=fill, size=9, wrap=True)
        ws.row_dimensions[row_idx].height = 28

    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
    else:
        for col_idx in range(1, ws.max_column + 1):
            vals = [str(ws.cell(row=r, column=col_idx).value or '')
                    for r in range(1, min(ws.max_row+1, 50))]
            ws.column_dimensions[get_column_letter(col_idx)].width = \
                min(max(len(v) for v in vals) + 2, 55)

# Build Overview sheet manually
ws_ov = wb['Overview']
overview_rows = [
    ('METRIC',                                         'FILE 1',        'FILE 2'),
    ('File name',                                       FILE1_PATH,      FILE2_PATH),
    ('Total rows',                                     f'{len(df1):,}', f'{len(df2):,}'),
    ('Total columns',                                  f'{len(df1.columns)-4}',
                                                       f'{len(df2.columns)-4}'),
    ('100% identical duplicate rows',                  f'{len(f1_full_dup):,}',
                                                       f'{len(f2_full_dup):,}'),
    ('Rows with same Name+Number+Year (dup groups)',   f'{len(f1_key3_dup):,}  →  {len(f1_dup_df)} groups',
                                                       f'{len(f2_key3_dup):,}  →  {len(f2_dup_df)} groups'),
    ('Rows with same Number+Year only (diff names)',   f'{len(f1_key2_dup):,}',
                                                       f'{len(f2_key2_dup):,}'),
    ('Rows with EXACT match in other file',            f'{int(f1_keys3_in_f2.sum()):,}',
                                                       f'{int(f2_keys3_in_f1.sum()):,}'),
    ('Rows with NO match in other file (Name+Num+Yr)', f'{int((~f1_keys3_in_f2).sum()):,}',
                                                       f'{int((~f2_keys3_in_f1).sum()):,}'),
    ('Truly new laws in File2 not in File1',           '—',
                                                       f'{len(extra_f2):,}'),
    ('end_date column exists',                         'NO',            'YES'),
    ('end_date non-null values',                       '—',
                                                       f'{int(df2["end_date"].notna().sum()):,} / {len(df2):,}'),
]

row_fills = {
    1: COLORS['header'],
    2: COLORS['blue'], 3: COLORS['blue'], 4: COLORS['blue'],
    5: COLORS['red'],  6: COLORS['amber'], 7: COLORS['amber'],
    8: COLORS['green'], 9: COLORS['amber'],
    10: COLORS['green'], 11: COLORS['blue'], 12: COLORS['blue'],
}

for r, row in enumerate(overview_rows, 1):
    fill = row_fills.get(r, COLORS['alt'] if r % 2 == 0 else COLORS['white'])
    for c, val in enumerate(row, 1):
        cell = ws_ov.cell(row=r, column=c, value=val)
        is_header_row = (r == 1)
        fmt(cell,
            fill=fill,
            bold=(r == 1 or c == 1),
            color='FFFFFFFF' if r == 1 else 'FF000000',
            align='center' if c > 1 else 'left',
            size=10 if r == 1 else 9)
    ws_ov.row_dimensions[r].height = 26

ws_ov.column_dimensions['A'].width = 48
ws_ov.column_dimensions['B'].width = 30
ws_ov.column_dimensions['C'].width = 30
ws_ov.freeze_panes = 'B2'

# Style all other sheets
style_sheet(wb['File1 Completeness'], col_widths=[30, 12, 12, 14])
style_sheet(wb['File2 Completeness'], col_widths=[30, 12, 12, 14])
style_sheet(wb['File1 Duplicates'],   col_widths=[55, 12, 8, 10, 20, 35, 22, 30])
style_sheet(wb['File2 Duplicates'],   col_widths=[55, 12, 8, 10, 35, 22, 30])
style_sheet(wb['Cross-File Analysis'],col_widths=[58, 14])
style_sheet(wb['Extra Laws in File2'],col_widths=[55, 12, 8, 16, 16, 16, 16])
style_sheet(wb['end_date Analysis'],  col_widths=[52, 40])

# Color the Recommendation column in duplicate sheets
for sheet_name in ['File1 Duplicates', 'File2 Duplicates']:
    ws_d = wb[sheet_name]
    rec_col = ws_d.max_column  # last column
    for row_idx in range(2, ws_d.max_row + 1):
        cell = ws_d.cell(row=row_idx, column=rec_col)
        val  = str(cell.value or '')
        fill = COLORS['amber'] if 'MERGE' in val else COLORS['green']
        fmt(cell, fill=fill, size=9, wrap=True)

wb.save(OUT_FILE)

# ──────────────────────────────────────────────────────────────────────────────
# PRINT SUMMARY TO TERMINAL
# ──────────────────────────────────────────────────────────────────────────────
print("[6/6] Done!\n")
print(f"  → {OUT_FILE}")
print()
print("  ╔══════════════════════════════════════════════════════╗")
print("  ║           PHASE 1 — EXPLORATION SUMMARY             ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  File1 rows              : {len(df1):>6,}                    ║")
print(f"  ║  File2 rows              : {len(df2):>6,}                    ║")
print(f"  ║  Difference              : {len(df2)-len(df1):>6,}                    ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  File1 — 100% dup rows   : {len(f1_full_dup):>6,}                    ║")
print(f"  ║  File1 — Name+Num+Yr dup : {len(f1_key3_dup):>6,} rows / {len(f1_dup_df):>2} groups    ║")
print(f"  ║  File1 — Num+Yr only dup : {len(f1_key2_dup):>6,}                    ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  File2 — 100% dup rows   : {len(f2_full_dup):>6,}                    ║")
print(f"  ║  File2 — Name+Num+Yr dup : {len(f2_key3_dup):>6,} rows / {len(f2_dup_df):>2} groups    ║")
print(f"  ║  File2 — Num+Yr only dup : {len(f2_key2_dup):>6,}                    ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  Exact matches F1↔F2     : {int(f1_keys3_in_f2.sum()):>6,}                    ║")
print(f"  ║  Truly new laws in File2 : {len(extra_f2):>6,}                    ║")
print(f"  ║  end_date non-null       : {int(df2['end_date'].notna().sum()):>6,} / {len(df2):,}         ║")
print("  ╚══════════════════════════════════════════════════════╝")
print()
print("  Open phase1_exploration_report.xlsx and share your observations!")
print("  Start with the 'Overview' sheet, then dig into each tab.")
