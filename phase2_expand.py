"""
Phase 2 — Enrich File1 with end_date + Append New Laws from File2
==================================================================
Cross-file matching strategy:
  Primary   : Num + Year + Magazine_Number  (high confidence)
  Fallback  : Num + Year only              (when Magazine missing)

Within-file duplicate detection (internal only):
  Norm_Name + Num + Year + Magazine_Number (all four — strictest)

Steps:
  1. Normalize Arabic text
  2. Match end_date from File2 → File1 (by Num+Year+Mag)
  3. Identify truly new laws in File2 (not in File1 by Num+Year+Mag)
  4. Append new laws to File1 with Source flag
  5. Export phase2_new_laws_preview.xlsx — review before accepting
  6. Export phase2_expanded.xlsx         — final combined file

Run: python phase2_expand.py
"""

import re, unicodedata
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
FILE1_PATH  = 'law_merged_output_V2.xlsx'
FILE2_PATH  = 'cleaned_v03_merged_updated.xlsx'
OUT_FINAL   = 'phase2_expanded.xlsx'
OUT_PREVIEW = 'phase2_new_laws_preview.xlsx'

# File2 col → File1 col
COL_MAP = {
    'leg_name'       : 'Law_Name',
    'leg_number'     : 'Law_Number',
    'year'           : 'Year',
    'magazine_number': 'Magazine_Number',
    'magazine_date'  : 'Magazine_Date',
    'magazine_page'  : 'Magazine_Page_Number',
    'start_date'     : 'Active_Date',
    'end_date'       : 'end_date',
    'replaced_for'   : 'Replaced_For',
    'status'         : 'Status',
}

COLORS = {
    'header'  : 'FF1F3864',
    'subhdr'  : 'FF2E75B6',
    'original': 'FFFFFFFF',
    'new_row' : 'FFDDEBF7',
    'end_date': 'FFC6EFCE',
    'alt'     : 'FFF2F7FF',
}

def thin_border():
    s = Side(style='thin', color='FFCCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def fmt(cell, fill=None, bold=False, color='FF000000', size=10,
        align='left', wrap=False):
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)
    cell.font      = Font(bold=bold, color=color, name='Arial', size=size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = thin_border()


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def normalize_arabic(text):
    if not isinstance(text, str) or not text.strip(): return ''
    text = ''.join(c for c in unicodedata.normalize('NFD', text)
                   if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[أإآٱ]','ا',text)
    text = re.sub(r'[ئؤ]','ء',text)
    text = text.replace('ة','ه').replace('ى','ي').replace('ـ','')
    text = re.sub(r'[().,،؛؟!«»]',' ',text)
    return re.sub(r'\s+',' ',text).strip()

def to_int_str(v):
    try: return str(int(float(v)))
    except: return ''

def clean_val(v):
    """Return None if value is empty/placeholder, else the string."""
    if v is None: return None
    s = str(v).strip()
    if s in ('','nan','0','NULL','None'): return None
    return s


# ─────────────────────────────────────────────────────────────────────────────
# 1. LOAD
# ─────────────────────────────────────────────────────────────────────────────
print("[1/7] Loading files ...")
df1 = pd.read_excel(FILE1_PATH)
df2 = pd.read_excel(FILE2_PATH)
print(f"      File1: {len(df1):,} rows")
print(f"      File2: {len(df2):,} rows")

# ── Cross-file keys (Num + Year + Mag) ───────────────────────────────────────
for df, num_col, year_col, mag_col in [
    (df1, 'Law_Number', 'Year',  'Magazine_Number'),
    (df2, 'leg_number', 'year',  'magazine_number'),
]:
    df['_num']  = df[num_col].apply(to_int_str)
    df['_year'] = df[year_col].apply(to_int_str)
    df['_mag']  = df[mag_col].apply(to_int_str)
    df['_key3'] = df['_num'] + '|' + df['_year'] + '|' + df['_mag']  # Num+Year+Mag
    df['_key2'] = df['_num'] + '|' + df['_year']                      # Num+Year only

# ── Within-file duplicate keys (Name + Num + Year + Mag) ─────────────────────
df1['_norm']  = df1['Law_Name'].apply(normalize_arabic)
df2['_norm']  = df2['leg_name'].apply(normalize_arabic)
df1['_kint']  = df1['_norm'] + '|' + df1['_num'] + '|' + df1['_year'] + '|' + df1['_mag']
df2['_kint']  = df2['_norm'] + '|' + df2['_num'] + '|' + df2['_year'] + '|' + df2['_mag']


# ─────────────────────────────────────────────────────────────────────────────
# 2. INTERNAL DUPLICATE REPORT (info only — no removal per your decision)
# ─────────────────────────────────────────────────────────────────────────────
f1_int_dups = df1[df1.duplicated(subset=['_kint'], keep=False)]
f2_int_dups = df2[df2.duplicated(subset=['_kint'], keep=False)]
print(f"\n      File1 internal dups (Name+Num+Year+Mag): {len(f1_int_dups):,} rows")
print(f"      File2 internal dups (Name+Num+Year+Mag): {len(f2_int_dups):,} rows")
print(f"      (Not removed — all columns must be 100% identical to remove)")


# ─────────────────────────────────────────────────────────────────────────────
# 3. MATCH end_date FROM FILE2 → FILE1
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2/7] Matching end_date from File2 to File1 ...")

# Build lookup dicts: key → end_date (prefer non-null values)
f2_key3_end = {}
f2_key2_end = {}

for _, row in df2.iterrows():
    v = clean_val(row['end_date'])
    if v:
        if row['_mag'] and row['_key3'] not in f2_key3_end:
            f2_key3_end[row['_key3']] = v
        if row['_key2'] not in f2_key2_end:
            f2_key2_end[row['_key2']] = v

def get_end_date(row):
    if row['_mag']:
        v = f2_key3_end.get(row['_key3'])
        if v: return v
    return f2_key2_end.get(row['_key2'], np.nan)

df1['end_date'] = df1.apply(get_end_date, axis=1)
end_filled = df1['end_date'].notna().sum()
print(f"      end_date filled for {end_filled:,} / {len(df1):,} existing File1 rows")


# ─────────────────────────────────────────────────────────────────────────────
# 4. IDENTIFY TRULY NEW LAWS
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3/7] Identifying truly new laws in File2 ...")

f1_key3_set = set(df1['_key3'])
f1_key2_set = set(df1['_key2'])

def is_new_law(row):
    """
    A law is NEW if it doesn't appear in File1.
    Uses Num+Year+Mag (primary) then Num+Year (fallback).
    """
    if row['_mag']:
        if row['_key3'] in f1_key3_set:
            return False
    if row['_key2'] in f1_key2_set:
        return False
    return True

df2['_is_new'] = df2.apply(is_new_law, axis=1)
new_laws       = df2[df2['_is_new']].copy()
matched_laws   = df2[~df2['_is_new']]

modern   = (new_laws['year'] >= 1950).sum()
historic = (new_laws['year'] <  1950).sum()

print(f"      New laws to append : {len(new_laws):,}")
print(f"      Already in File1   : {len(matched_laws):,}")
print(f"      Modern  (≥1950)    : {modern:,}")
print(f"      Historic (<1950)   : {historic:,}")


# ─────────────────────────────────────────────────────────────────────────────
# 5. BUILD NEW ROWS
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4/7] Building new rows ...")

orig_cols = [c for c in df1.columns if not c.startswith('_')]
df1['Source'] = 'Original'

new_rows = []
for _, row2 in new_laws.iterrows():
    new_row = {col: np.nan for col in orig_cols}
    for f2_col, f1_col in COL_MAP.items():
        if f2_col in row2.index:
            v = clean_val(row2[f2_col])
            if v:
                new_row[f1_col] = v
    new_row['Source'] = 'Added from File2'
    new_rows.append(new_row)

new_rows_df = pd.DataFrame(new_rows)
print(f"      New rows ready: {len(new_rows_df):,}")


# ─────────────────────────────────────────────────────────────────────────────
# 6. COMBINE
# ─────────────────────────────────────────────────────────────────────────────
print("\n[5/7] Combining ...")

all_cols = list(orig_cols) + ['Source']
for c in new_rows_df.columns:
    if c not in all_cols:
        all_cols.append(c)

for c in all_cols:
    if c not in df1.columns:         df1[c]          = np.nan
    if c not in new_rows_df.columns: new_rows_df[c]  = np.nan

df_final = pd.concat([df1[all_cols], new_rows_df[all_cols]], ignore_index=True)
final_save = df_final[[c for c in all_cols if not c.startswith('_')]].copy()

print(f"      Final rows : {len(final_save):,}")
print(f"      Columns    : {len(final_save.columns)}")


# ─────────────────────────────────────────────────────────────────────────────
# 7. EXPORT & FORMAT
# ─────────────────────────────────────────────────────────────────────────────
print("\n[6/7] Exporting ...")

# Preview (new laws only)
preview_cols = ['Law_Name','Law_Number','Year','Magazine_Number',
                'Magazine_Date','Active_Date','end_date',
                'Replaced_For','Status','Source']
preview_df = new_rows_df[[c for c in preview_cols if c in new_rows_df.columns]].copy()
preview_df.to_excel(OUT_PREVIEW, index=False)

# Full file
final_save.to_excel(OUT_FINAL, index=False)

col_widths = {
    'Law_Name':45,'Law_Number':12,'Year':8,'Magazine_Number':16,
    'Magazine_Date':16,'Magazine_Page_Number':18,'Active_Date':16,
    'end_date':16,'Replaced_For':35,'Status':14,'Source':20,
    'FullName':40,'pmk_ID':8,'LobID':10,
}

print("[7/7] Formatting ...")
for out_path, is_preview in [(OUT_FINAL, False), (OUT_PREVIEW, True)]:
    wb  = load_workbook(out_path)
    ws  = wb.active
    ws.freeze_panes = 'A2'
    ws.title = 'New Laws Preview' if is_preview else 'Laws — Expanded'

    cols_ws = [c.value for c in ws[1]]
    enddate_idx = cols_ws.index('end_date') + 1 if 'end_date' in cols_ws else None
    orig_count  = len(df1)

    # Header
    for ci, cn in enumerate(cols_ws, 1):
        cell = ws.cell(row=1, column=ci)
        hf = COLORS['subhdr'] if cn in ('end_date','Source') else COLORS['header']
        fmt(cell, fill=hf, bold=True, color='FFFFFFFF', align='center', size=10)
    ws.row_dimensions[1].height = 36

    # Data rows
    for ri in range(2, ws.max_row + 1):
        is_new = is_preview or (ri > orig_count + 1)
        for ci, cn in enumerate(cols_ws, 1):
            cell = ws.cell(row=ci, column=ci)
            cell = ws.cell(row=ri, column=ci)
            if is_new:
                fill = COLORS['new_row']
            elif cn == 'end_date' and cell.value not in (None,'','nan'):
                fill = COLORS['end_date']
            elif ri % 2 == 0:
                fill = COLORS['alt']
            else:
                fill = COLORS['original']
            fmt(cell, fill=fill, size=9,
                bold=(is_new and cn in ('Law_Name','Source')))
        ws.row_dimensions[ri].height = 22

    # Widths
    for ci, cn in enumerate(cols_ws, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(cn, 13)

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print()
print("  ╔══════════════════════════════════════════════════════╗")
print("  ║              PHASE 2 — SUMMARY                      ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  Original File1 rows          : {len(df1):>6,}               ║")
print(f"  ║  New laws appended from File2 : {len(new_rows_df):>6,}               ║")
print(f"  ║  Final total rows             : {len(final_save):>6,}               ║")
print("  ╠══════════════════════════════════════════════════════╣")
print(f"  ║  end_date filled (existing)   : {end_filled:>6,}               ║")
print(f"  ║  New modern laws  (≥1950)     : {modern:>6,}               ║")
print(f"  ║  New historic laws (<1950)    : {historic:>6,}               ║")
print("  ╠══════════════════════════════════════════════════════╣")
print("  ║  New columns added:                                  ║")
print("  ║    → end_date  (from File2, matched by Num+Year+Mag) ║")
print("  ║    → Source    (Original / Added from File2)         ║")
print("  ╚══════════════════════════════════════════════════════╝")
print()
print(f"  ⚠️  REVIEW FIRST  → {OUT_PREVIEW}")
print(f"  ✅  Full result   → {OUT_FINAL}")
print()
print("  Color guide:")
print("    🟦 Light blue rows  = new laws added from File2")
print("    🟩 Light green cell = end_date filled for existing law")
print("    ⬜ White/gray rows  = original File1 records")
